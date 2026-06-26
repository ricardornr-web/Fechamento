"""
ml_api.py — Integração OAuth + Orders com a API do Mercado Livre.

Funções públicas:
  get_auth_url(client_id, redirect_uri, state) -> str
  exchange_code(client_id, client_secret, code, redirect_uri) -> dict
  refresh_access_token(client_id, client_secret, refresh_tok) -> dict
  get_user_id(access_token) -> int
  fetch_orders(access_token, seller_id, date_from, date_to) -> list
  orders_to_excel_bytes(orders, empresa) -> bytes
"""

import base64
import hashlib
import io
import secrets
from datetime import datetime

import requests
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------

OAUTH_URL = "https://auth.mercadolivre.com.br/authorization"
TOKEN_URL = "https://api.mercadolibre.com/oauth/token"
API_BASE  = "https://api.mercadolibre.com"

COR_MAE   = "FFB7B7B7"
COR_FILHO = "FFF3F3F3"

STATUS_PT = {
    "paid":               "Pagamento confirmado",
    "cancelled":          "Cancelado",
    "pending":            "Pendente",
    "in_process":         "Em processo",
    "partially_refunded": "Reembolso parcial",
    "refunded":           "Reembolso total",
}

# Colunas do relatório ML — na mesma ordem esperada por ml_core.py
HEADERS_ML = [
    "N.º de venda",
    "Data da venda",
    "Estado",
    "Descrição do status",
    "Comprador",
    "SKU",
    "# de anúncio",
    "Canal de venda",
    "Título do anúncio",
    "Variação",
    "Preço unitário de venda do anúncio (BRL)",
    "Tipo de anúncio",
    "Unidades",
    "Receita por produtos (BRL)",
    "Receita por acréscimo no preço (pago pelo comprador)",
    "Taxa de parcelamento equivalente ao acréscimo",
    "Tarifa de venda e impostos (BRL)",
    "Receita por envio (BRL)",
    "Tarifas de envio (BRL)",
    "Custo de envio com base nas medidas e peso declarados",
    "Custo por diferenças nas medidas e no peso do pacote",
    "Descontos e bônus",
    "Cancelamentos e reembolsos (BRL)",
    "Total (BRL)",
    "Loja oficial",
]

# ---------------------------------------------------------------------------
# OAuth
# ---------------------------------------------------------------------------

def generate_pkce() -> tuple:
    """Gera (code_verifier, code_challenge) para o fluxo PKCE."""
    verifier  = secrets.token_urlsafe(43)
    digest    = hashlib.sha256(verifier.encode()).digest()
    challenge = base64.urlsafe_b64encode(digest).rstrip(b"=").decode()
    return verifier, challenge


def get_auth_url(client_id: str, redirect_uri: str, state: str = "",
                 code_challenge: str = "") -> str:
    url = (
        f"{OAUTH_URL}?response_type=code"
        f"&client_id={client_id}"
        f"&redirect_uri={redirect_uri}"
        f"&scope=offline_access"
    )
    if state:
        url += f"&state={state}"
    if code_challenge:
        url += f"&code_challenge={code_challenge}&code_challenge_method=S256"
    return url


def exchange_code(client_id: str, client_secret: str, code: str,
                  redirect_uri: str, code_verifier: str = "") -> dict:
    data = {
        "grant_type":    "authorization_code",
        "client_id":     client_id,
        "client_secret": client_secret,
        "code":          code,
        "redirect_uri":  redirect_uri,
    }
    if code_verifier:
        data["code_verifier"] = code_verifier
    resp = requests.post(TOKEN_URL, data=data)
    if not resp.ok:
        raise Exception(f"ML token error {resp.status_code}: {resp.text}")
    return resp.json()


def refresh_access_token(client_id: str, client_secret: str, refresh_tok: str) -> dict:
    resp = requests.post(TOKEN_URL, data={
        "grant_type":    "refresh_token",
        "client_id":     client_id,
        "client_secret": client_secret,
        "refresh_token": refresh_tok,
    })
    if not resp.ok:
        raise Exception(f"ML refresh error {resp.status_code}: {resp.text}")
    return resp.json()


def get_user_info(access_token: str) -> dict:
    resp = requests.get(
        f"{API_BASE}/users/me",
        headers={"Authorization": f"Bearer {access_token}"},
    )
    resp.raise_for_status()
    data = resp.json()
    return {"id": data["id"], "nickname": data.get("nickname", "")}


def get_user_id(access_token: str) -> int:
    return get_user_info(access_token)["id"]


# ---------------------------------------------------------------------------
# Busca de pedidos
# ---------------------------------------------------------------------------

def fetch_orders(access_token: str, seller_id: int, date_from: str, date_to: str) -> list:
    """
    date_from / date_to: strings ISO8601, ex. "2024-01-01T00:00:00.000-03:00"
    Retorna lista de dicts de pedidos da API do ML.
    """
    headers = {"Authorization": f"Bearer {access_token}"}
    orders  = []
    offset  = 0
    limit   = 50

    while True:
        url = (
            f"{API_BASE}/orders/search"
            f"?seller={seller_id}"
            f"&order.status=all"
            f"&date_closed.from={date_from}"
            f"&date_closed.to={date_to}"
            f"&sort=date_asc"
            f"&limit={limit}&offset={offset}"
        )
        resp = requests.get(url, headers=headers)
        resp.raise_for_status()
        data    = resp.json()
        results = data.get("results", [])
        if not results:
            break
        orders.extend(results)
        total = data.get("paging", {}).get("total", 0)
        offset += len(results)
        if offset >= total:
            break

    return orders


# ---------------------------------------------------------------------------
# Conversão para Excel no formato esperado por ml_core.py
# ---------------------------------------------------------------------------

def _fmt_date(iso_str: str) -> str:
    if not iso_str:
        return ""
    try:
        dt = datetime.fromisoformat(iso_str.replace("Z", "+00:00"))
        return dt.strftime("%d/%m/%Y %H:%M:%S")
    except Exception:
        return iso_str


def _variacao_str(attrs: list) -> str:
    if not attrs:
        return ""
    return ", ".join(
        f"{v.get('name', '')}: {v.get('value_name', '')}"
        for v in attrs
        if v.get("value_name")
    )


def orders_to_excel_bytes(orders: list, empresa: str) -> bytes:
    """
    Converte pedidos da API do ML em bytes de Excel no formato que ml_core.py espera.
    Pedidos com 1 item → linha simples (sem cor).
    Pedidos com N itens → linha mãe (cor FFB7B7B7) + N linhas filho (cor FFF3F3F3).
    """
    fill_mae   = PatternFill("solid", start_color=COR_MAE)
    fill_filho = PatternFill("solid", start_color=COR_FILHO)
    n_cols     = len(HEADERS_ML)
    idx        = {h: i + 1 for i, h in enumerate(HEADERS_ML)}

    wb = Workbook()
    ws = wb.active

    # Cabeçalho na linha 1 (ml_core busca "N.º de venda" nas linhas 1-14, coluna 1)
    for ci, h in enumerate(HEADERS_ML, 1):
        ws.cell(row=1, column=ci, value=h)

    row_num = 2

    for order in orders:
        order_id  = str(order.get("id", ""))
        status    = order.get("status", "")
        status_pt = STATUS_PT.get(status, status)
        date_str  = _fmt_date(order.get("date_closed") or order.get("date_created", ""))
        buyer     = (order.get("buyer") or {}).get("nickname", "")
        items     = order.get("order_items", [])
        payments  = order.get("payments", [])

        if not items:
            continue

        ship_rev = round(sum(float(p.get("shipping_cost") or 0) for p in payments), 2)
        tot_fee  = round(sum(float(it.get("sale_fee") or 0) for it in items), 2)
        tot_rev  = round(
            sum(float(it.get("unit_price") or 0) * int(it.get("quantity") or 0) for it in items), 2
        )

        def _common(rn: int):
            ws.cell(rn, idx["N.º de venda"],   value=order_id)
            ws.cell(rn, idx["Data da venda"],  value=date_str)
            ws.cell(rn, idx["Estado"],         value=status_pt)
            ws.cell(rn, idx["Comprador"],      value=buyer)
            ws.cell(rn, idx["Canal de venda"], value="Mercado Livre")
            ws.cell(rn, idx["Loja oficial"],   value=empresa)

        if len(items) == 1:
            it         = items[0]
            item_info  = it.get("item") or {}
            unit_price = float(it.get("unit_price") or 0)
            quantity   = int(it.get("quantity") or 0)
            sale_fee   = float(it.get("sale_fee") or 0)
            rec_prod   = round(unit_price * quantity, 2)

            _common(row_num)
            ws.cell(row_num, idx["SKU"],               value=item_info.get("seller_sku", ""))
            ws.cell(row_num, idx["# de anúncio"],      value=str(item_info.get("id", "")))
            ws.cell(row_num, idx["Título do anúncio"], value=item_info.get("title", ""))
            ws.cell(row_num, idx["Variação"],          value=_variacao_str(it.get("variation_attributes")))
            ws.cell(row_num, idx["Preço unitário de venda do anúncio (BRL)"], value=unit_price)
            ws.cell(row_num, idx["Tipo de anúncio"],   value=it.get("listing_type_id", ""))
            ws.cell(row_num, idx["Unidades"],          value=quantity)
            ws.cell(row_num, idx["Receita por produtos (BRL)"],           value=rec_prod)
            ws.cell(row_num, idx["Tarifa de venda e impostos (BRL)"],     value=-sale_fee)
            ws.cell(row_num, idx["Receita por envio (BRL)"],              value=ship_rev)
            ws.cell(row_num, idx["Total (BRL)"],       value=round(rec_prod - sale_fee + ship_rev, 2))
            row_num += 1

        else:
            # Linha mãe
            _common(row_num)
            ws.cell(row_num, idx["Unidades"],                             value=sum(int(it.get("quantity") or 0) for it in items))
            ws.cell(row_num, idx["Receita por produtos (BRL)"],           value=tot_rev)
            ws.cell(row_num, idx["Tarifa de venda e impostos (BRL)"],     value=-tot_fee)
            ws.cell(row_num, idx["Receita por envio (BRL)"],              value=ship_rev)
            ws.cell(row_num, idx["Total (BRL)"],                          value=round(tot_rev - tot_fee + ship_rev, 2))
            for ci in range(1, n_cols + 1):
                ws.cell(row_num, ci).fill = fill_mae
            row_num += 1

            # Linhas filho
            for it in items:
                item_info  = it.get("item") or {}
                unit_price = float(it.get("unit_price") or 0)
                quantity   = int(it.get("quantity") or 0)
                rec_prod   = round(unit_price * quantity, 2)

                ws.cell(row_num, idx["N.º de venda"],      value=order_id)
                ws.cell(row_num, idx["Canal de venda"],    value="Mercado Livre")
                ws.cell(row_num, idx["SKU"],               value=item_info.get("seller_sku", ""))
                ws.cell(row_num, idx["# de anúncio"],      value=str(item_info.get("id", "")))
                ws.cell(row_num, idx["Título do anúncio"], value=item_info.get("title", ""))
                ws.cell(row_num, idx["Variação"],          value=_variacao_str(it.get("variation_attributes")))
                ws.cell(row_num, idx["Preço unitário de venda do anúncio (BRL)"], value=unit_price)
                ws.cell(row_num, idx["Tipo de anúncio"],   value=it.get("listing_type_id", ""))
                ws.cell(row_num, idx["Unidades"],          value=quantity)
                ws.cell(row_num, idx["Receita por produtos (BRL)"], value=rec_prod)
                for ci in range(1, n_cols + 1):
                    ws.cell(row_num, ci).fill = fill_filho
                row_num += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
