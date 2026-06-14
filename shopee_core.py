# shopee_core.py
# ==============
# Versão do consolidar_shopee.py adaptada para rodar em memória (Streamlit).
# Não lê/grava arquivos em disco — recebe bytes e retorna bytes.
#
# Interface pública:
#   processar(arquivos, tabela_bytes) -> (xlsx_bytes: bytes, logs: list[str])
#
# Dependências: openpyxl, pandas

import io
import re
import math
import unicodedata
from datetime import datetime
from collections import defaultdict

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# CONSTANTES
# =============================================================================

CHAVE_RICAPET = "order_creation_date"

COL_ID_PEDIDO   = "ID do pedido"
COL_STATUS      = "Status do pedido"
COL_SKU_PRINC   = "Nº de referência do SKU principal"
COL_NOME_PROD   = "Nome do Produto"
COL_SKU         = "Número de referência SKU"
COL_VARIACAO    = "Nome da variação"
COL_PRECO_ORI   = "Preço original"
COL_PRECO_ACORD = "Preço acordado"
COL_QUANTIDADE  = "Quantidade"
COL_SUBTOTAL    = "Subtotal do produto"
COL_VALOR_TOTAL = "Valor Total"
COL_TAXA_ENVIO  = "Taxa de envio pagas pelo comprador"
COL_COM_BRUTA   = "Taxa de comissão bruta"
COL_SERV_BRUTA  = "Taxa de serviço bruta"
COL_TOTAL_GLOB  = "Total global"

TOLERANCIA = 0.02

COLUNAS_DISTRIBUIR = [
    "Taxa de transação",
    "Taxa de comissão líquida",
    "Taxa de serviço líquida",
    "Taxa de Envio Reversa",
    "Cupom do vendedor",
    "Coin Cashback Voucher Amount Sponsored by Seller",
    "Cupom",
    "Incentivo de cupom",
    "Ajuste por pagamento via PIX",
    "Incentivo Shopee para ação comercial",
    "Ajuste por participação em ação comercial",
    "Desconto Shopee da Leve Mais por Menos",
    "Desconto da Leve Mais por Menos do vendedor",
    "Compensar Moedas Shopee",
    "Total descontado Cartão de Crédito",
    "Total global",
    "Valor estimado do frete",
]

COLUNAS_PRODUTO = {
    COL_ID_PEDIDO, COL_STATUS, "Hot Listing", "Cancelar Motivo",
    "Status da Devolução / Reembolso", "Número de rastreamento",
    "Opção de envio", "Método de envio",
    "Data de criação do pedido", "Hora do pagamento do pedido",
    "Data prevista de envio", "Tempo de Envio", "Domestic Delivered Date",
    "Hora completa do pedido", "Data da Finalização do Cancelamento",
    "Pedido FBS", COL_SKU_PRINC, COL_NOME_PROD, COL_SKU, COL_VARIACAO,
    "Shopee Owned", COL_PRECO_ORI, COL_PRECO_ACORD, COL_QUANTIDADE,
    "Returned quantity", COL_SUBTOTAL,
    "Desconto do vendedor", "Desconto do vendedor_2",
    "Peso total SKU", "Número de produtos pedidos", "Peso total do pedido",
    "Código do Cupom", "Indicador da Leve Mais por Menos",
    "Nome de usuário (comprador)", "Nome do destinatário",
    "Telefone", "CPF do Comprador", "Endereço de entrega",
    "Cidade", "Bairro", "UF", "País", "CEP",
    "Observação do comprador", "Nota",
    "Desconto de Frete Aproximado",
    "_empresa", "Empresa", "_cor",
}

# =============================================================================
# LOG SIMPLES
# =============================================================================

class SimpleLog:
    def __init__(self):
        self.msgs = []

    def info(self, msg):
        self.msgs.append(f"ℹ️  {msg}")

    def warning(self, msg):
        self.msgs.append(f"⚠️  {msg}")

    def error(self, msg):
        self.msgs.append(f"❌  {msg}")

# =============================================================================
# UTILITÁRIOS
# =============================================================================

def sf(v, default=0.0):
    if v is None:
        return default
    if isinstance(v, float) and math.isnan(v):
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default

def vazio(v):
    if v is None:
        return True
    if isinstance(v, float) and math.isnan(v):
        return True
    return str(v).strip() in ("", " ", "nan")

def arr(v):
    return round(v, 2)

def _norm(s):
    s = str(s).strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.lower().split())

def _find_col(colunas, *candidatos):
    mapa = {_norm(c): c for c in colunas}
    for cand in candidatos:
        real = mapa.get(_norm(cand))
        if real is not None:
            return real
    return None

# =============================================================================
# IDENTIFICAÇÃO DE EMPRESA
# =============================================================================

def identificar_empresa(nome: str):
    n = nome.lower()
    if not n.startswith("order.all."):
        return None
    if CHAVE_RICAPET in n:
        return "Ricapet"
    return "Thapets"

# =============================================================================
# LEITURA DOS RELATÓRIOS
# =============================================================================

def ler_arquivo(conteudo: bytes, nome: str, empresa: str, log):
    log.info(f"Lendo {nome}  [{empresa}]")
    wb = load_workbook(io.BytesIO(conteudo), data_only=True)
    ws = wb.active

    raw_headers = []
    seen = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        if not h:
            h = f"_col_{c}"
        h = str(h).strip()
        seen[h] = seen.get(h, 0) + 1
        if seen[h] > 1:
            h = f"{h}_{seen[h]}"
        raw_headers.append(h)

    rows = []
    for r in range(2, ws.max_row + 1):
        v1 = ws.cell(row=r, column=1).value
        if v1 is None:
            continue
        row = {"_empresa": empresa, "Empresa": empresa}
        for c, h in enumerate(raw_headers, 1):
            row[h] = ws.cell(row=r, column=c).value
        rows.append(row)

    wb.close()
    log.info(f"  → {len(rows)} linhas  |  {len(raw_headers)} colunas")
    return rows, raw_headers

# =============================================================================
# UNIFICAÇÃO
# =============================================================================

def unificar(ricapet_rows, ricapet_cols, thapets_rows, thapets_cols, log):
    extra    = [c for c in thapets_cols if c not in ricapet_cols]
    all_cols = ["Empresa"] + ricapet_cols + extra

    def norm(rows, base_cols, extra_cols):
        out = []
        for r in rows:
            row = {"Empresa": r.get("Empresa", r.get("_empresa", "")),
                   "_empresa": r.get("_empresa", "")}
            for c in base_cols + extra_cols:
                row[c] = r.get(c)
            out.append(row)
        return out

    unified    = norm(ricapet_rows, ricapet_cols, extra) + norm(thapets_rows, thapets_cols, extra)
    ricapet_ct = sum(1 for r in unified if r["Empresa"] == "Ricapet")
    thapets_ct = sum(1 for r in unified if r["Empresa"] == "Thapets")
    log.info(f"Unificado: {len(unified)} linhas  |  {len(all_cols)} colunas")
    log.info(f"  Ricapet: {ricapet_ct}  |  Thapets: {thapets_ct}")
    return unified, all_cols

# =============================================================================
# DISTRIBUIÇÃO PROPORCIONAL
# =============================================================================

def distribuir_valores(rows, all_cols, log):
    col_sub_real = _find_col(all_cols, COL_SUBTOTAL, "Subtotal do produto")
    col_id_real  = _find_col(all_cols, COL_ID_PEDIDO, "ID do pedido")
    col_env_real = _find_col(all_cols, COL_TAXA_ENVIO, "Taxa de envio pagas pelo comprador")

    if not col_sub_real or not col_id_real:
        log.warning("Colunas de ID pedido ou Subtotal não encontradas; distribuição ignorada")
        return rows

    cols_para_dist = []
    for cand in COLUNAS_DISTRIBUIR:
        real = _find_col(all_cols, cand)
        if real and real not in COLUNAS_PRODUTO:
            cols_para_dist.append(real)

    for row in rows:
        row["_taxa_envio_original"] = sf(row.get(col_env_real))

    grupos = defaultdict(list)
    for i, row in enumerate(rows):
        pid = str(row.get(col_id_real, "")).strip()
        grupos[pid].append(i)

    multi_count = sum(1 for idxs in grupos.values() if len(idxs) > 1)
    log.info(f"Pedidos únicos: {len(grupos)}  |  Com múltiplos produtos: {multi_count}")

    for pid, idxs in grupos.items():
        if len(idxs) == 1:
            continue

        subtotais = [sf(rows[i].get(col_sub_real)) for i in idxs]
        total_sub = sum(subtotais)

        if total_sub == 0:
            pcts = [1.0 / len(idxs)] * len(idxs)
        else:
            pcts = [s / total_sub for s in subtotais]

        for col in cols_para_dist:
            val_original = sf(rows[idxs[0]].get(col))
            valores = [arr(val_original * p) for p in pcts]
            diff = arr(val_original - arr(sum(valores)))
            valores[-1] = arr(valores[-1] + diff)
            for i, v in zip(idxs, valores):
                rows[i][col] = v

    return rows

# =============================================================================
# CÁLCULO DO TOTAL (BRL) LÍQUIDO
# =============================================================================

def calcular_total_liquido(rows, all_cols, log):
    col_vt   = _find_col(all_cols, COL_VALOR_TOTAL,  "Valor Total")
    col_env  = _find_col(all_cols, COL_TAXA_ENVIO,   "Taxa de envio pagas pelo comprador")
    col_com  = _find_col(all_cols, COL_COM_BRUTA,    "Taxa de comissão bruta")
    col_serv = _find_col(all_cols, COL_SERV_BRUTA,   "Taxa de serviço bruta")
    col_st   = _find_col(all_cols, COL_STATUS,       "Status do pedido")

    if not col_vt:
        log.error("Coluna 'Valor Total' não encontrada; Total (BRL) não calculado")
        for row in rows:
            row["Total (BRL)"] = None
        return rows

    ajustes_thapets = 0
    for row in rows:
        st = str(row.get(col_st, "") or "").strip().lower() if col_st else ""
        if "não pago" in st or "nao pago" in st or "order received" in st:
            row["Total (BRL)"] = 0
            continue

        vt   = sf(row.get(col_vt))
        env  = sf(row.get(col_env))  if col_env  else 0.0
        com  = sf(row.get(col_com))  if col_com  else 0.0
        serv = sf(row.get(col_serv)) if col_serv else 0.0

        net = arr(vt - env - com - serv)

        env_orig = sf(row.get("_taxa_envio_original"))
        if (row.get("Empresa") == "Thapets"
                and abs(env_orig - 8.0) < 0.001
                and "conclu" in st):
            net = arr(net + 8.0)
            ajustes_thapets += 1

        row["Total (BRL)"] = net

    if ajustes_thapets:
        log.info(f"  Thapets ajuste +R$8: {ajustes_thapets} linha(s)")

    log.info(f"Total (BRL) calculado para {len(rows)} linhas")
    return rows

# =============================================================================
# CARREGAMENTO DA TABELA AUXILIAR
# =============================================================================

def carregar_auxiliares(tabela_bytes: bytes, log):
    xl = pd.ExcelFile(io.BytesIO(tabela_bytes))

    def norm_sheet(name):
        n = unicodedata.normalize("NFKD", str(name))
        n = "".join(c for c in n if not unicodedata.combining(c))
        return n.strip().upper()

    sheet_map = {norm_sheet(s): s for s in xl.sheet_names}
    log.info(f"  Abas TABELA_AUXILIAR: {xl.sheet_names}")

    chave_tp = norm_sheet("TABELA_PRODUTOS")
    if chave_tp not in sheet_map:
        raise ValueError(f"Aba 'TABELA_PRODUTOS' não encontrada. Abas: {xl.sheet_names}")
    tp = xl.parse(sheet_map[chave_tp], dtype=str).fillna("")
    tp.columns = [str(c).strip() for c in tp.columns]
    col_map_tp = {str(c).strip().upper(): c for c in tp.columns}
    for col in ["SKU", "TÍTULO", "PRODUTO", "COR", "TAMANHO", "UNIDADE"]:
        if col not in col_map_tp:
            raise ValueError(f"Coluna '{col}' não encontrada em TABELA_PRODUTOS. "
                             f"Colunas: {list(tp.columns)}")
    rename_tp = {}
    for canon, upper in [("SKU","SKU"), ("Título","TÍTULO"), ("PRODUTO","PRODUTO"),
                          ("COR","COR"), ("TAMANHO","TAMANHO"), ("UNIDADE","UNIDADE")]:
        real = col_map_tp[upper]
        if real != canon:
            rename_tp[real] = canon
    if rename_tp:
        tp = tp.rename(columns=rename_tp)

    chave_st = norm_sheet("STATUS")
    if chave_st not in sheet_map:
        raise ValueError(f"Aba 'STATUS' não encontrada. Abas: {xl.sheet_names}")
    st = xl.parse(sheet_map[chave_st], dtype=str).fillna("")
    st.columns = [str(c).strip() for c in st.columns]
    col_map_st = {str(c).strip().upper(): c for c in st.columns}
    col_status = col_map_st.get("STATUS")
    col_manter = col_map_st.get("MANTER")
    if not col_status or not col_manter:
        raise ValueError(f"Aba STATUS precisa das colunas 'Status' e 'Manter'. "
                         f"Encontradas: {list(st.columns)}")
    rename_st = {}
    if col_status != "Status": rename_st[col_status] = "Status"
    if col_manter != "Manter": rename_st[col_manter] = "Manter"
    if rename_st:
        st = st.rename(columns=rename_st)

    chave_custo = norm_sheet("CUSTO")
    tabela_custo_fech = pd.DataFrame()
    if chave_custo in sheet_map:
        tc = xl.parse(sheet_map[chave_custo]).fillna("")
        tc.columns = [str(c).strip() for c in tc.columns]
        cm = {str(c).strip().upper(): c for c in tc.columns}
        c_prod = cm.get("PRODUTO"); c_cor = cm.get("COR")
        c_tam  = cm.get("TAMANHO"); c_cst = cm.get("CUSTO")
        if c_prod and c_cst:
            ren = {}
            if c_prod != "PRODUTO":              ren[c_prod] = "PRODUTO"
            if c_cor  and c_cor  != "COR":       ren[c_cor]  = "COR"
            if c_tam  and c_tam  != "TAMANHO":   ren[c_tam]  = "TAMANHO"
            if c_cst  != "CUSTO":                ren[c_cst]  = "CUSTO"
            if ren: tc = tc.rename(columns=ren)
            for col in ["COR", "TAMANHO"]:
                if col not in tc.columns: tc[col] = ""
            tc["PRODUTO"]  = tc["PRODUTO"].astype(str).str.strip()
            tc["COR"]      = tc["COR"].astype(str).str.strip()
            tc["TAMANHO"]  = tc["TAMANHO"].astype(str).str.strip()
            tc["CUSTO"]    = pd.to_numeric(tc["CUSTO"], errors="coerce")
            tc = tc[tc["PRODUTO"].str.strip() != ""].dropna(subset=["CUSTO"])
            tabela_custo_fech = tc[["PRODUTO","COR","TAMANHO","CUSTO"]].reset_index(drop=True)
            log.info(f"  Aba CUSTO: {len(tabela_custo_fech)} registros")
    else:
        log.warning("  Aba CUSTO não encontrada em TABELA_AUXILIAR.xlsx")

    tabela_custo_unit = pd.DataFrame(columns=["PRODUTO", "TAMANHO", "CUSTO_UNITARIO"])
    for nome_aba in ["CUSTO_UNITARIO", "Planilha2"]:
        chave_u = norm_sheet(nome_aba)
        if chave_u not in sheet_map:
            continue
        raw = xl.parse(sheet_map[chave_u], header=None)
        hdr = None
        for i_r, row_r in raw.iterrows():
            if any(str(v).strip().upper() == "PRODUTO" for v in row_r if pd.notna(v)):
                hdr = i_r; break
        if hdr is None:
            continue
        raw.columns = [str(v).strip() for v in raw.iloc[hdr]]
        raw = raw.iloc[hdr + 1:].reset_index(drop=True)
        cm = {str(c).strip().upper(): c for c in raw.columns}
        c_p = cm.get("PRODUTO")
        c_t = cm.get("TAMANHO")
        c_c = next((v for k,v in cm.items() if "CUSTO" in k and "UNIT" in k),
                   next((v for k,v in cm.items() if "CUSTO" in k), None))
        if not c_p or not c_c:
            continue
        ren = {}
        if c_p != "PRODUTO":         ren[c_p] = "PRODUTO"
        if c_t and c_t != "TAMANHO": ren[c_t] = "TAMANHO"
        if c_c != "CUSTO_UNITARIO":  ren[c_c] = "CUSTO_UNITARIO"
        if ren: raw = raw.rename(columns=ren)
        if "TAMANHO" not in raw.columns: raw["TAMANHO"] = ""
        raw["PRODUTO"]        = raw["PRODUTO"].astype(str).str.strip()
        raw["TAMANHO"]        = raw["TAMANHO"].astype(str).str.strip().replace("nan","")
        raw["CUSTO_UNITARIO"] = pd.to_numeric(raw["CUSTO_UNITARIO"], errors="coerce")
        raw = raw[~raw["PRODUTO"].str.upper().str.contains("TOTAL", na=False)]
        raw = raw[raw["PRODUTO"].str.strip() != ""].dropna(subset=["CUSTO_UNITARIO"])
        tabela_custo_unit = raw[["PRODUTO","TAMANHO","CUSTO_UNITARIO"]].reset_index(drop=True)
        log.info(f"  Aba {nome_aba}: {len(tabela_custo_unit)} registros de custo unitário")
        break

    log.info(f"TABELA_AUXILIAR: {len(tp)} produtos  |  {len(st)} status  |  "
             f"{len(tabela_custo_fech)} custos Fechamento")
    return tp, st, tabela_custo_fech, tabela_custo_unit

# =============================================================================
# CRIAÇÃO DA BASE FINAL
# =============================================================================

def criar_base_final(rows, all_cols, log):
    data_cols = [c for c in all_cols if c not in ("Empresa", "_empresa")]
    base_cols = ["Empresa"] + data_cols

    registros = []
    for row in rows:
        reg = {}
        for col in base_cols:
            if col == "Empresa":
                reg[col] = row.get("Empresa", row.get("_empresa", ""))
            else:
                val = row.get(col)
                reg[col] = None if (isinstance(val, float) and math.isnan(val)) else val
        reg["Total (BRL)"] = row.get("Total (BRL)")
        registros.append(reg)

    df = pd.DataFrame(registros)

    if "Total (BRL)" not in df.columns:
        df["Total (BRL)"] = None

    log.info(f"Base Final: {len(df)} linhas")
    return df

# =============================================================================
# PARSE DE VARIAÇÃO
# =============================================================================

_CORES_CONHECIDAS = {
    "preto","branco","cinza","grafite","azul","vermelho","verde",
    "amarelo","laranja","roxo","rosa","marrom","bege","caramelo",
    "chumbo","sisal","creme","nude","natural","azul marinho","off white",
}
_TAM_RE    = re.compile(r"^\d+[xX]\d+$")
_ROUPA_RE  = re.compile(r"^(P|M|G|GG|XGG)$", re.IGNORECASE)
_UNID_RE   = re.compile(r"\d+\s*[Uu]nid", re.IGNORECASE)
_UNID_RE2  = re.compile(r"^\d+\s*[Uu]nidade", re.IGNORECASE)

def _parse_variacao(variacao) -> dict:
    if vazio(variacao):
        return {"COR": "", "TAMANHO": ""}
    parts = [p.strip() for p in str(variacao).split(",")]
    cor = ""; tam = ""
    for part in parts:
        p_low = _norm(part)
        if _UNID_RE.search(p_low) or _UNID_RE2.match(p_low):
            continue
        if _TAM_RE.match(part.strip()) or _ROUPA_RE.match(part.strip()):
            if not tam:
                tam = part.strip()
        elif p_low in _CORES_CONHECIDAS:
            if not cor:
                cor = part.strip()
        else:
            sub = part.strip().split()
            for s in sub:
                if _TAM_RE.match(s) and not tam:
                    tam = s
                elif _norm(s) in _CORES_CONHECIDAS and not cor:
                    cor = s
            if not cor and not _TAM_RE.match(part.strip()):
                cor = part.strip()
    return {"COR": cor, "TAMANHO": tam}

# =============================================================================
# PREENCHIMENTO DAS COLUNAS CALCULADAS
# =============================================================================

def preencher_colunas(df, tabela_produtos, tabela_status, log, tabela_custo_unit=None):
    for col in ["PRODUTO", "COR", "TAMANHO", "UNIDADE", "MANTER"]:
        df[col] = ""
    df["Custo"]  = None
    df["Alerta"] = ""

    campos_str = ["PRODUTO", "COR", "TAMANHO", "UNIDADE"]

    _col_custo_tp = next(
        (c for c in tabela_produtos.columns if str(c).strip().upper() == "CUSTO"), None
    )
    if _col_custo_tp is None:
        log.warning("Coluna 'CUSTO' não encontrada em TABELA_PRODUTOS")

    lookup_sku    = {}
    lookup_titulo = {}

    for _, row in tabela_produtos.iterrows():
        sku    = str(row.get("SKU",    "")).strip()
        titulo = str(row.get("Título", "")).strip()
        dados  = {c: str(row.get(c, "")).strip() for c in campos_str}
        _raw_c = row.get(_col_custo_tp, "") if _col_custo_tp else ""
        _raw_s = str(_raw_c).strip()
        dados["CUSTO"] = None
        if _raw_s not in ("", "nan", "NaN", "None", "-"):
            try:
                dados["CUSTO"] = float(_raw_s)
            except (ValueError, TypeError):
                pass

        if sku   and sku.lower()   not in lookup_sku:
            lookup_sku[sku.lower()]     = dados
        if titulo and titulo.lower() not in lookup_titulo:
            lookup_titulo[titulo.lower()] = dados

    lookup_status = {
        str(row["Status"]).strip().lower(): str(row["Manter"]).strip()
        for _, row in tabela_status.iterrows()
        if str(row.get("Status", "")).strip()
    }

    col_sku    = _find_col(df.columns, COL_SKU,      "Número de referência SKU",  "SKU")
    col_titulo = _find_col(df.columns, COL_NOME_PROD, "Nome do Produto", "Título do anúncio")
    col_estado = _find_col(df.columns, COL_STATUS,    "Status do pedido", "Estado")
    col_var    = _find_col(df.columns, COL_VARIACAO,  "Nome da variação")

    if col_sku is None:
        log.warning("Coluna 'Número de referência SKU' não encontrada — cruzamento por SKU desativado")

    acertos_sku = acertos_titulo = acertos_status = 0

    for idx in df.index:
        sku    = str(df.at[idx, col_sku]).strip()    if col_sku    else ""
        titulo = str(df.at[idx, col_titulo]).strip() if col_titulo else ""
        estado = str(df.at[idx, col_estado]).strip() if col_estado else ""
        var    = df.at[idx, col_var]                 if col_var    else None

        dados = lookup_sku.get(sku.lower()) if sku else None
        if dados:
            for c in campos_str:
                df.at[idx, c] = dados[c]
            if dados["CUSTO"] is not None:
                df.at[idx, "Custo"] = dados["CUSTO"]
            acertos_sku += 1
        else:
            dados = lookup_titulo.get(titulo.lower()) if titulo else None
            if dados:
                for c in campos_str:
                    df.at[idx, c] = dados[c]
                if dados["CUSTO"] is not None:
                    df.at[idx, "Custo"] = dados["CUSTO"]
                acertos_titulo += 1
            else:
                pv = _parse_variacao(var)
                if pv["COR"]:
                    df.at[idx, "COR"] = pv["COR"]
                if pv["TAMANHO"]:
                    df.at[idx, "TAMANHO"] = pv["TAMANHO"]

        if estado:
            manter = lookup_status.get(estado.lower(), "")
            df.at[idx, "MANTER"] = manter
            if manter:
                acertos_status += 1

    if tabela_custo_unit is not None and not tabela_custo_unit.empty:
        lk_pt = {}
        lk_p  = {}
        for _, rc in tabela_custo_unit.iterrows():
            pk = _norm(rc.get("PRODUTO", ""))
            tk = _norm(rc.get("TAMANHO", ""))
            cv = rc.get("CUSTO_UNITARIO")
            if pk and cv is not None and not (isinstance(cv, float) and cv != cv):
                lk_pt[(pk, tk)] = float(cv)
                if tk in ("", "-", "—"):
                    lk_p[pk] = float(cv)

        col_prod_bf = _find_col(df.columns, "PRODUTO")
        col_tam_bf  = _find_col(df.columns, "TAMANHO")
        if col_prod_bf:
            acertos_pt = 0
            for idx in df.index:
                if df.at[idx, "Custo"] is not None:
                    continue
                pk = _norm(df.at[idx, col_prod_bf])
                tk = _norm(df.at[idx, col_tam_bf]) if col_tam_bf else ""
                cv = (lk_pt.get((pk, tk)) or lk_pt.get((pk, "")) or lk_p.get(pk))
                if cv is not None:
                    df.at[idx, "Custo"] = cv
                    acertos_pt += 1
            log.info(f"  Custo via PRODUTO+TAMANHO: {acertos_pt} registros")

    _kit_re = re.compile(r'KIT(\d+)', re.IGNORECASE)
    if col_sku:
        for idx in df.index:
            sku_val = str(df.at[idx, col_sku]).strip()
            m = _kit_re.search(sku_val)
            if m:
                df.at[idx, "UNIDADE"] = m.group(1)

    col_total_bf = _find_col(df.columns, "Total (BRL)")
    if col_total_bf:
        total_num = pd.to_numeric(df[col_total_bf], errors="coerce")
        df["Alerta"] = total_num.apply(
            lambda v: "⚠" if (v is not None and not (isinstance(v, float) and v != v) and v <= 0) else ""
        )
        n_alertas = int((df["Alerta"] == "⚠").sum())
        if n_alertas:
            log.warning(f"  {n_alertas} linha(s) com Total (BRL) <= 0 marcadas com ⚠")

    col_qtd  = _find_col(df.columns, COL_QUANTIDADE, "Quantidade")
    col_unid = _find_col(df.columns, "UNIDADE")
    if col_qtd and col_unid:
        qtd_n  = pd.to_numeric(df[col_qtd],  errors="coerce").fillna(0)
        unid_n = pd.to_numeric(df[col_unid], errors="coerce").fillna(1)
        df["TOTAL UNIDADES"] = (qtd_n * unid_n).round(0).astype(int)
        if "Alerta" in df.columns:
            df.loc[df["Alerta"] == "⚠", "TOTAL UNIDADES"] = 0
    else:
        df["TOTAL UNIDADES"] = None
        log.warning("TOTAL UNIDADES não calculado: colunas ausentes")

    acertos_custo = int(df["Custo"].notna().sum())
    log.info(f"Cruzamento SKU: {acertos_sku}  |  Título: {acertos_titulo}  |  "
             f"MANTER: {acertos_status}  |  Custo: {acertos_custo}")
    return df

# =============================================================================
# ABA BASE FINAL
# =============================================================================

def _hdr(ws, row, n_cols, bg="1F4E79", fg="FFFFFF"):
    fill  = PatternFill("solid", start_color=bg)
    font  = Font(name="Calibri", bold=True, color=fg, size=11)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin  = Side(style="thin", color="BFBFBF")
    brd   = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font; cell.fill = fill
        cell.alignment = align; cell.border = brd
    ws.row_dimensions[row].height = 30

def _escrever_base_final(wb, base_final_df):
    ws = wb.create_sheet("Base Final")
    df = base_final_df.copy()

    COLS_FRONT = ["Alerta", "Empresa", "PRODUTO", "COR", "TAMANHO",
                  "UNIDADE", "MANTER", "Custo", "TOTAL UNIDADES", "Total (BRL)"]
    todas      = list(df.columns)
    internas   = {"_empresa", "_taxa_envio_original", "Tem"}
    colunas_bf = (
        [c for c in COLS_FRONT if c in todas]
        + [c for c in todas if c not in COLS_FRONT and c not in internas]
    )
    df = df[[c for c in colunas_bf if c in df.columns]].copy()

    _thin  = Side(style="thin",  color="E0E0E0")
    _BRD   = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _LA    = Alignment(horizontal="left",   vertical="center")
    _CA    = Alignment(horizontal="center", vertical="center")
    _RA    = Alignment(horizontal="right",  vertical="center")

    _HDR_ORIG = "404040"
    _HDR_CALC = "1F4E79"
    _HDR_ALRT = "C00000"
    _ROW_ALRT = PatternFill("solid", start_color="FFE5E5")

    COLS_CALC = {"PRODUTO","COR","TAMANHO","UNIDADE","MANTER","Custo",
                 "TOTAL UNIDADES","Total (BRL)"}
    ws.row_dimensions[1].height = 20
    colunas_reais = list(df.columns)

    for ci, h in enumerate(colunas_reais, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        if h == "Alerta":
            bg_hdr = _HDR_ALRT
        elif h in COLS_CALC or h == "Empresa":
            bg_hdr = _HDR_CALC
        else:
            bg_hdr = _HDR_ORIG
        cell.fill      = PatternFill("solid", start_color=bg_hdr)
        cell.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
        cell.alignment = _CA
        cell.border    = _BRD

    col_alerta_idx = (colunas_reais.index("Alerta") if "Alerta" in colunas_reais else None)

    for ri, row_data in enumerate(df.itertuples(index=False), start=2):
        row_list   = list(row_data)
        tem_alerta = (str(row_list[col_alerta_idx]).strip() == "⚠"
                      if col_alerta_idx is not None else False)
        row_fill   = _ROW_ALRT if tem_alerta else None

        for ci, (col, valor) in enumerate(zip(colunas_reais, row_list), start=1):
            val  = None if (isinstance(valor, float) and math.isnan(valor)) else valor
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(
                name="Calibri", size=9,
                bold  = (col == "Alerta" and tem_alerta),
                color = "C00000" if (col == "Alerta" and tem_alerta) else "1A1A1A",
            )
            if row_fill:
                cell.fill = row_fill
            cell.border = _BRD

            if col == "Alerta":
                cell.alignment = _CA
            elif isinstance(val, (int, float)):
                cell.alignment     = _RA
                cell.number_format = "#,##0.00"
            else:
                cell.alignment = _LA
        ws.row_dimensions[ri].height = 14

    n_front = len([c for c in COLS_FRONT if c in colunas_reais])
    ws.freeze_panes = f"{get_column_letter(n_front + 1)}2"
    ws.auto_filter.ref = ws.dimensions

    LARG = {
        "Alerta":7, "Empresa":10, "PRODUTO":24, "COR":13, "TAMANHO":11,
        "UNIDADE":9, "MANTER":9, "Custo":11, "TOTAL UNIDADES":11, "Total (BRL)":13,
        COL_ID_PEDIDO:22, COL_STATUS:16, COL_SKU:32, COL_NOME_PROD:38,
        COL_VARIACAO:20, COL_QUANTIDADE:9, COL_SUBTOTAL:14,
        COL_VALOR_TOTAL:14, COL_TAXA_ENVIO:18, COL_COM_BRUTA:16, COL_SERV_BRUTA:16,
    }
    for ci, col in enumerate(colunas_reais, 1):
        ws.column_dimensions[get_column_letter(ci)].width = LARG.get(col, 14)

# =============================================================================
# ABA FECHAMENTO
# =============================================================================

def _escrever_fechamento(wb, base_final_df, tabela_custo_fech=None):
    ws = wb.create_sheet("Fechamento")
    df = base_final_df.copy()

    def _col(df, *names):
        for n in names:
            m = next((x for x in df.columns if str(x).strip().lower() == n.lower()), None)
            if m: return m
        return None

    col_p  = _col(df, "PRODUTO")
    col_c  = _col(df, "COR")
    col_t  = _col(df, "TAMANHO")
    col_r  = _col(df, "Total (BRL)")
    col_qu = _col(df, "TOTAL UNIDADES") or _col(df, "Quantidade")

    def _n(c):
        return pd.to_numeric(df[c], errors="coerce").fillna(0) if c else pd.Series([0.0]*len(df))

    df["_rec"] = _n(col_r)
    df["_qty"] = _n(col_qu)

    _ca = next((c for c in df.columns if str(c).strip() == "Alerta"), None)
    if _ca: df = df[df[_ca].astype(str).str.strip() != "⚠"].copy()
    if col_p: df = df[df[col_p].astype(str).str.strip() != ""].copy()

    def _norm_f(s):
        s = str(s).strip()
        s = unicodedata.normalize("NFKD", s)
        s = "".join(ch for ch in s if not unicodedata.combining(ch))
        return " ".join(s.lower().split())

    _SEM = {"","-","—","unica","única","unidade","único","unico","nan","none"}
    _ALIAS = {
        _norm_f("Pega Areia"):              [_norm_f("Tapete pega areia")],
        _norm_f("Arranhador de Braço"):     [_norm_f("Arranhador Braço")],
        _norm_f("Arranhador de Braço Simples"): [_norm_f("Arranhador Braço Simples")],
        _norm_f("Arranhador de Madeira"):   [_norm_f("Arranhador Madeira")],
        _norm_f("Brinquedo Interativo"):    [_norm_f("Bririnquedo interativo")],
        _norm_f("Arranhador Túnel"):        [_norm_f("Tunel para gatos")],
        _norm_f("Sanitario e Pega Areia"):  [_norm_f("Sanitário + PA"),_norm_f("Sanitario + PA")],
        _norm_f("Alimentador Automático"):  [_norm_f("Comedouro Automatico"),_norm_f("Alimentador Automatico")],
        _norm_f("Banheiro Inteligente"):    [_norm_f("Sanitário Automatico"),_norm_f("Sanitario Automatico")],
        _norm_f("Pato Interativo"):         [_norm_f("Pato Ducky")],
        _norm_f("Piso Carpete"):            [_norm_f("Piso Carpete 50x50")],
        _norm_f("Arranhador Cara de Gato"): [_norm_f("Arranhador cara gato"),_norm_f("Arranhador Adesivo Gato")],
        _norm_f("Camisa Dryfit"):           [_norm_f("Camisa Dry Fit")],
        _norm_f("Cama Suspensa"):           [_norm_f("Caminha Rede")],
        _norm_f("Poste Arranhador"):        [_norm_f("Arranhador Poste")],
    }

    _lk = {}
    if tabela_custo_fech is not None and not tabela_custo_fech.empty:
        for _, _r in tabela_custo_fech.iterrows():
            _pk = _norm_f(_r["PRODUTO"]); _ck = _norm_f(_r.get("COR",""))
            _tk = _norm_f(_r.get("TAMANHO","")); _v  = _r["CUSTO"]
            if _pk and pd.notna(_v) and float(_v) > 0:
                _ck2 = "" if _ck in _SEM else _ck; _tk2 = "" if _tk in _SEM else _tk
                _lk[(_pk, _ck2, _tk2)] = float(_v)

    def _custo(prod, cor="", tam=""):
        pk = _norm_f(prod); ck = _norm_f(cor); ck = "" if ck in _SEM else ck
        tk = _norm_f(tam);  tk = "" if tk in _SEM else tk
        for cand in [pk] + _ALIAS.get(pk, []):
            v = _lk.get((cand, ck, tk))
            if v: return v
        return 0.0

    _SEM_FILTRO = {"","-","—","unica","única","nan","none"}

    def _soma(prod, col_f1=None, v1=None, col_f2=None, v2=None):
        mask = (df[col_p].astype(str).str.strip().str.lower() == str(prod).strip().lower()
                if col_p else pd.Series([True]*len(df)))
        if col_f1 and v1 is not None and col_f1 in df.columns:
            if str(v1).strip().lower() not in _SEM_FILTRO:
                mask &= df[col_f1].astype(str).str.strip().str.lower() == str(v1).strip().lower()
        if col_f2 and v2 is not None and col_f2 in df.columns:
            if str(v2).strip().lower() not in _SEM_FILTRO:
                mask &= df[col_f2].astype(str).str.strip().str.lower() == str(v2).strip().lower()
        sub  = df[mask]
        qty  = sub["_qty"].sum(); rec = sub["_rec"].sum()
        cor_v = v1 if col_f1 == col_c else (v2 if col_f2 == col_c else "")
        tam_v = v1 if col_f1 == col_t else (v2 if col_f2 == col_t else "")
        return qty, rec, _custo(prod, cor=cor_v or "", tam=tam_v or "")

    NOBRD  = Border()
    L  = Alignment(horizontal="left",   vertical="center")
    C  = Alignment(horizontal="center", vertical="center")
    R  = Alignment(horizontal="right",  vertical="center")
    FMT_M = 'R$ #,##0.00'; FMT_Z = 'R$ #,##0.00_);"-"'
    FMT_I = '#,##0'; FMT_P = '0%'
    CIANO = "00B0F0"; AMARELO = "FFFF00"

    def _f(bold=False, color="000000", size=11):
        return Font(name="Calibri", bold=bold, color=color, size=size)
    def _bg(h): return PatternFill("solid", start_color=h)
    fn = _f()

    for cl, wd in [("A",8.89),("B",22.33),("C",56.33),("D",8.55),("E",12.44),
                   ("F",14.66),("G",12.78),("H",12.89),("I",15.78),("J",14.0),("K",4.44)]:
        ws.column_dimensions[cl].width = wd

    def _w(row, col, val, font=None, fill=None, align=None, fmt=None):
        cell = ws.cell(row=row, column=col, value=val)
        cell.border = NOBRD
        if font:  cell.font  = font
        if fill:  cell.fill  = fill
        if align: cell.alignment = align
        if fmt:   cell.number_format = fmt
        return cell

    _t = Side(style="thin",   color="BFBFBF")
    _m = Side(style="medium", color="595959")

    def _dado(lin, val_b, val_c, qty, rec, custo, is_first=False, is_last=False, sep_tam=False):
        _sep = Side(style="thin", color="595959")
        top = _m if is_first else _t
        bot = _m if is_last  else (_sep if sep_tam else _t)
        brd_bc = Border(top=top, bottom=bot, left=_m, right=Side())
        brd_cc = Border(top=top, bottom=bot, left=Side(), right=Side())
        brd_dk = Border(top=top, bottom=bot, left=_t, right=_t)
        brd_k  = Border(top=top, bottom=bot, left=_t, right=_m)

        cell_b = ws.cell(row=lin, column=2)
        if val_b is not None:
            cell_b.value = val_b; cell_b.font = fn; cell_b.alignment = R
        cell_b.border = brd_bc

        cell_c = ws.cell(row=lin, column=3)
        cell_c.value = str(val_c) if val_c is not None else ""
        cell_c.font = fn; cell_c.alignment = L; cell_c.border = brd_cc

        _w(lin,4,int(qty),fn,align=R,fmt=FMT_I);               ws.cell(lin,4).border=brd_dk
        _w(lin,5,f"=IFERROR(F{lin}/D{lin},0)",fn,align=R,fmt=FMT_M); ws.cell(lin,5).border=brd_dk
        _w(lin,6,round(rec,2) if rec else 0,fn,align=R,fmt=FMT_Z);   ws.cell(lin,6).border=brd_dk
        _w(lin,7,round(custo,2) if custo else 0,fn,align=R,fmt=FMT_Z);ws.cell(lin,7).border=brd_dk
        _w(lin,8,f"=G{lin}*D{lin}",fn,align=R,fmt=FMT_M);     ws.cell(lin,8).border=brd_dk
        _w(lin,9,f"=IFERROR(E{lin}-G{lin},0)",fn,align=R,fmt=FMT_Z); ws.cell(lin,9).border=brd_dk
        _w(lin,10,f"=F{lin}-H{lin}",fn,align=R,fmt=FMT_M);    ws.cell(lin,10).border=brd_dk
        _w(lin,11,f'=IFERROR(IF(F{lin}=0,"",J{lin}/F{lin}),"")',fn,align=R,fmt=FMT_P)
        ws.cell(lin,11).border=brd_k
        ws.row_dimensions[lin].height = 15

    def _total(lin, r0, r1):
        ws.row_dimensions[lin].height = 15
        brd_tot_b  = Border(bottom=_m, left=_m, right=Side())
        brd_tot_c  = Border(bottom=_m, left=Side(), right=Side())
        brd_tot_dk = Border(top=_t, bottom=_m, left=_t, right=_t)
        brd_tot_k  = Border(top=_t, bottom=_m, left=_t, right=_m)
        ws.cell(lin,2).border = brd_tot_b
        ws.cell(lin,3).border = brd_tot_c
        _w(lin,4,f"=SUM(D{r0}:D{r1})",fn,align=R,fmt=FMT_I);  ws.cell(lin,4).border=brd_tot_dk
        ws.cell(lin,5).border = brd_tot_dk
        _w(lin,6,f"=SUM(F{r0}:F{r1})",fn,fill=_bg(CIANO),align=R,fmt=FMT_M); ws.cell(lin,6).border=brd_tot_dk
        ws.cell(lin,7).border = brd_tot_dk
        _w(lin,8,f"=SUM(H{r0}:H{r1})",fn,align=R,fmt=FMT_M);  ws.cell(lin,8).border=brd_tot_dk
        ws.cell(lin,9).border = brd_tot_dk
        _w(lin,10,f"=SUM(J{r0}:J{r1})",fn,fill=_bg(AMARELO),align=R,fmt=FMT_M); ws.cell(lin,10).border=brd_tot_dk
        _w(lin,11,f'=IFERROR(IF(F{lin}=0,"",J{lin}/F{lin}),"")',fn,align=R,fmt=FMT_P)
        ws.cell(lin,11).border = brd_tot_k

    _TAM_ORD = {"P":1,"M":2,"G":3,"GG":4,"XGG":5}
    def _sort_tam(t):
        t = str(t).strip()
        if t in _TAM_ORD: return (0, _TAM_ORD[t], t)
        try:
            parts = t.replace("x","X").split("X")
            return (1, int(parts[0]), t)
        except: return (2, 0, t)

    ESTRUTURA = [
        ("Pega Areia","Cor","COR",["Preto","Azul","Amarelo","Rosa"],"Pega Areia"),
        ("Tapete Lavável","Modelo","TAM",["P","M","G"],"Tapete Lavável"),
        ("Tapete Comedouro","Cor","COR",["Amarelo","Azul","Laranja","Marrom","Roxo","Verde","Vermelho"],"Tapete Comedouro"),
        ("Arranhador de Braço","Modelo","TAM_COR_MAP",
         [("70X50","P", ["Azul","Bege","Cinza"]),("100X50","M",["Azul","Bege","Cinza"]),
          ("120X50","G",["Azul","Bege","Cinza"])],"Arranhador de Braço"),
        ("Arranhador de Braço Simples","Modelo","TAM_COR_MAP",
         [("70X50","P", ["Azul","Bege","Cinza"]),("100X50","M",["Azul","Bege","Cinza"]),
          ("120X50","G",["Azul","Bege","Cinza"])],"Arranhador de Braço Simples"),
        ("Arranhador braço 1 folha","Modelo","TAM_COR_MAP",
         [("70X50","P", ["Azul","Bege","Cinza"]),("100X50","M",["Azul","Bege","Cinza"]),
          ("120X50","G",["Azul","Bege","Cinza"])],"Arranhador Carpete 1 Folha"),
        ("Arranhador Adesivo","Modelo","TAM_COR",
         [("50x30",["Cinza","Grafite","Preto","Marrom","Bege","Azul"]),
          ("70x50",["Cinza","Grafite","Preto","Marrom","Bege","Azul"]),
          ("100x50",["Cinza","Grafite","Preto","Marrom","Bege","Azul"]),
          ("120x50",["Cinza","Grafite","Preto","Marrom","Bege","Azul"]),
          ("200x50",["Cinza","Grafite","Preto","Marrom","Bege","Azul"])],None),
        ("Arranhador de Madeira","Modelo","COR",["60cm","Cinza","Bege","Azul"],"Arranhador de Madeira"),
        ("Casinha Toca","Cor","COR",["Bege","Cinza","Caramelo","Verde","Azul","Chumbo"],None),
        ("Capa para Carro","Cor","COR",["Preto"],None),
        ("Casinha e Arranhador","Cor","COR",["Único"],"Casinha e Arranhador"),
        ("Camisa Dryfit","Modelo","TAM_COR",
         [("P", ["Preto","Cinza","Verde","Branco"]),("M", ["Preto","Cinza","Verde","Branco"]),
          ("G", ["Preto","Cinza","Verde","Branco"]),("GG",["Preto","Cinza","Verde","Branco"]),
          ("XGG",["Preto","Cinza","Verde","Branco"])],"Camisa Dryfit"),
        ("Casinhas / Caminhas","Modelo","GRUPO_B",
         [("Casinha 01","Bege"),("Casinha 01","Cinza"),("Casinha 01","Azul"),
          ("Caminha 02","Bege"),("Caminha 02","Cinza"),("Caminha 02","Azul"),
          ("Casinha 03","Bege"),("Casinha 03","Cinza"),("Casinha 03","Azul"),
          ("Casinha 04","Bege"),("Casinha 04","Cinza"),("Casinha 04","Azul"),
          ("Casinha 05","Bege"),("Casinha 05","Cinza"),("Casinha 05","Azul"),
          ("Casinha 06","Bege"),("Casinha 06","Cinza"),("Casinha 06","Azul"),
          ("Casinha 07","Bege"),("Casinha 07","Cinza"),("Casinha 07","Azul"),
          ("Casinha 08","Bege"),("Casinha 08","Cinza"),("Casinha 08","Azul"),
          ("Casinha 09","Bege"),("Casinha 09","Cinza"),("Casinha 09","Azul"),
          ("Casinha 10","Bege"),("Casinha 10","Cinza"),("Casinha 10","Azul"),
          ("Casinha 11","Bege"),("Casinha 11","Cinza"),("Casinha 11","Azul"),
          ("Casinha 12","Bege"),("Casinha 12","Cinza"),("Casinha 12","Azul"),
          ("Casinha 13","Bege"),("Casinha 13","Cinza"),("Casinha 13","Azul"),
          ("Casinha 14","Bege"),("Casinha 14","Cinza"),("Casinha 14","Azul"),
          ("Casinha 15","Bege"),("Casinha 15","Cinza"),("Casinha 15","Azul"),
          ("Casinha 16","Bege"),("Casinha 16","Cinza"),("Casinha 16","Azul")],None),
        ("Playground","Modelo","COR",["-"],"Playground"),
        ("Chuchups","Modelo","COR",["-"],None),
        ("Presilhas Aspirais","Modelo","COR",["-"],None),
        ("Piso Carpete","Modelo","COR",["Cinza","Grafite","Preto","Marrom","Bege","Azul"],"Piso Carpete"),
        ("Sanitário","Modelo","COR",["Bege","Cinza","Preto","Verde"],"Sanitário"),
        ("Sanitario e Pega Areia","Cor","COR",["Bege","Cinza","Preto","Verde"],"Sanitario e Pega Areia"),
        ("Cama Suspensa","Modelo","COR",["Bege","Cinza","Caramelo","Verde","Azul","Chumbo"],"Cama Suspensa"),
        ("Arranhador Túnel","Modelo","COR",["Bege","Cinza","Caramelo","Verde","Azul","Chumbo"],"Arranhador Túnel"),
        ("Degrau","Modelo","COR",["Bege","Cinza","Grafite","Marrom","Verde","Azul","Chumbo"],None),
        ("Caminha Comeia","Modelo","COR",["Bege","Cinza","Caramelo","Verde","Azul","Chumbo"],None),
        ("Brinquedo Interativo","Modelo","COR",["Azul","Bege","Cinza","Grafite","Marrom","Preto"],"Brinquedo Interativo"),
        ("Arranhadores","Cor","GRUPO_B",
         [("Arranhador Rampa","Bege"),("Arranhador Rampa","Cinza"),("Arranhador Rampa","Azul"),
          ("Arranhador Parede","Sisal"),("Arranhador Parede","Bege"),("Arranhador Parede","Cinza"),
          ("Arranhador Parede","Preto"),("Arranhador Parede","Marrom"),
          ("Arranhador Parede","Azul"),("Arranhador Parede","Grafite"),
          ("Arranhador Cara de Gato","Bege"),("Arranhador Cara de Gato","Cinza"),
          ("Arranhador Cara de Gato","Preto"),("Arranhador Cara de Gato","Verde"),
          ("Arranhador Cara de Gato","Azul"),("Arranhador Cara de Gato","Chumbo"),
          ("Poste Arranhador","Sisal")],None),
        ("Comedouro automatico","Cor","COR",["Branco"],"Alimentador Automático"),
        ("Rampa 6mm","Cor","COR",["Azul","Bege","Cinza"],"Rampa 6mm"),
        ("Arranhador Adesivo Gato","Modelo","COR",["Azul","Bege","Cinza","Grafite","Marrom","Preto"],None),
        ("Pato Interativo","Modelo","COR",["Amarelo"],"Pato Interativo"),
        ("Banheiro Inteligente","Modelo","COR",["-"],"Banheiro Inteligente"),
        ("Comedouro Inteligente","Modelo","COR",["Branco"],None),
        ("Playgrounds 15MM","Modelo","GRUPO_B",
         [("Playground 02","-"),("Playground 03","-"),("Playground 04","-"),
          ("Playground 05","-"),("Playground 06","-")],None),
        ("Nicho Grande","Modelo","COR",["Única"],None),
        ("FullSellers","Modelo","COR",
         ["Bolsa Caixa Transporte Aves Calopsita Periquito Pássaros Top Cor Azul",
          "Gaiola Hamster Tubo Super Luxo 3 Andares Porquinho Da India",
          "Gaiola Hamster 2 Andares Porquinho Da Inda Chinchila"],None),
    ]

    ws.row_dimensions[3].height = 15
    _w(3, 2, "SHOPEE", fn, fill=_bg(AMARELO), align=L)
    linha = 4; linhas_total = []

    for nome_exib, tipo_exib, modo, itens, nome_bf in ESTRUTURA:
        prod_bf = nome_bf if nome_bf else nome_exib
        ws.row_dimensions[linha].height = 15
        brd_hdr_b  = Border(top=_m, bottom=_t, left=_m, right=Side())
        brd_hdr_c  = Border(top=_m, bottom=_t, left=Side(), right=Side())
        brd_hdr_dk = Border(top=_m, bottom=_t, left=_t, right=_t)
        brd_hdr_k  = Border(top=_m, bottom=_t, left=_t, right=_m)
        _w(linha,2,nome_exib,fn,align=L); ws.cell(linha,2).border=brd_hdr_b
        _w(linha,3,tipo_exib,fn,align=L); ws.cell(linha,3).border=brd_hdr_c
        for col_n, lbl in [(4,"Unidades"),(5,"Valor Unitario"),(6,"Valor de entrada"),
                           (7,"Custo Unitario"),(8,"Custo Total"),
                           (9,"Lucro por unidade"),(10,"Lucro total")]:
            brd = brd_hdr_k if col_n == 10 else brd_hdr_dk
            _w(linha,col_n,lbl,fn,align=C); ws.cell(linha,col_n).border=brd
        ws.cell(linha,11).border = Border(top=_m, bottom=_t, left=_t, right=_m)
        linha += 1; r_ini = linha

        if modo == "COR":
            itens_ord = sorted(itens, key=str.lower); n_itens = len(itens_ord)
            for i_cor, cor in enumerate(itens_ord):
                qty, rec, cu = _soma(prod_bf, col_c, cor)
                _dado(linha,None,cor,qty,rec,cu,is_first=(i_cor==0),is_last=(i_cor==n_itens-1))
                linha += 1
        elif modo == "TAM":
            n_itens = len(itens)
            for i_tam, tam in enumerate(itens):
                qty, rec, cu = _soma(prod_bf, col_t, tam)
                _dado(linha,None,tam,qty,rec,cu,is_first=(i_tam==0),is_last=(i_tam==n_itens-1))
                linha += 1
        elif modo == "TAM_COR":
            todas_linhas = [(tam,cor) for tam,cores in itens for cor in sorted(cores,key=str.lower)]
            n_total = len(todas_linhas); i_global = 0
            for i_tam, (tam, cores) in enumerate(itens):
                cores_ord = sorted(cores, key=str.lower); ultimo_tam = (i_tam == len(itens)-1)
                for i_cor, cor in enumerate(cores_ord):
                    ultima_cor = (i_cor == len(cores_ord)-1)
                    is_last_block = (i_global == n_total-1); is_sep_tam = ultima_cor and not ultimo_tam
                    qty, rec, cu = _soma(prod_bf, col_t, tam, col_c, cor)
                    _dado(linha,tam,cor,qty,rec,cu,is_first=(i_global==0),is_last=is_last_block,sep_tam=is_sep_tam)
                    i_global += 1; linha += 1
        elif modo == "TAM_COR_MAP":
            todas_linhas = [(r,t,cor) for r,t,cores in itens for cor in sorted(cores,key=str.lower)]
            n_total = len(todas_linhas); i_global = 0
            for i_tam, (rotulo, tam_bf, cores) in enumerate(itens):
                cores_ord = sorted(cores, key=str.lower); ultimo_tam = (i_tam == len(itens)-1)
                for i_cor, cor in enumerate(cores_ord):
                    ultima_cor = (i_cor == len(cores_ord)-1)
                    is_last_block = (i_global == n_total-1); is_sep_tam = ultima_cor and not ultimo_tam
                    qty, rec, cu = _soma(prod_bf, col_t, tam_bf, col_c, cor)
                    if cu == 0: cu = _custo(prod_bf, cor=cor, tam=rotulo)
                    _dado(linha,rotulo,cor,qty,rec,cu,is_first=(i_global==0),is_last=is_last_block,sep_tam=is_sep_tam)
                    i_global += 1; linha += 1
        elif modo == "GRUPO_B":
            itens_validos = [(sb,sc) for sb,sc in itens if sb]; n_itens = len(itens_validos)
            for i_item, (sub_b, sub_c) in enumerate(itens_validos):
                sub_prod = str(sub_b)
                if sub_c: qty, rec, cu = _soma(sub_prod, col_c, sub_c)
                else: qty, rec, cu = _soma(sub_prod); cu = _custo(sub_prod)
                _dado(linha,sub_b,sub_c,qty,rec,cu,is_first=(i_item==0),is_last=(i_item==n_itens-1))
                linha += 1

        r_fim = linha - 1
        _total(linha, r_ini, r_fim)
        linhas_total.append(linha); linha += 1
        ws.row_dimensions[linha].height = 10; linha += 1

    ws.row_dimensions[linha].height = 7; linha += 1
    ws.row_dimensions[linha].height = 15
    refs_d = "+".join(f"D{r}" for r in linhas_total)
    refs_f = "+".join(f"F{r}" for r in linhas_total)
    refs_h = "+".join(f"H{r}" for r in linhas_total)
    refs_j = "+".join(f"J{r}" for r in linhas_total)
    _w(linha,2,"TOTAL",fn,align=L)
    _w(linha,4,f"={refs_d}",fn,align=R,fmt=FMT_I)
    _w(linha,6,f"={refs_f}",fn,align=R,fmt=FMT_M)
    _w(linha,8,f"={refs_h}",fn,align=R,fmt=FMT_M)
    _w(linha,10,f"={refs_j}",fn,align=R,fmt=FMT_M)
    _w(linha,11,f'=IFERROR(IF(F{linha}=0,"",J{linha}/F{linha}),"")',fn,align=R,fmt=FMT_P)
    ws.freeze_panes = "A4"

# =============================================================================
# ABA PRODUTOS
# =============================================================================

def _escrever_painel_produtos(wb, base_final_df):
    ws  = wb.create_sheet("Produtos")
    df  = base_final_df.copy()

    def _f(bold=False, color="1A1A1A", size=10, italic=False):
        return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)
    def _bg(h): return PatternFill("solid", start_color=h)
    thin = Side(style="thin", color="D9D9D9")
    BRD  = Border(left=thin, right=thin, top=thin, bottom=thin)
    L = Alignment(horizontal="left",   vertical="center")
    C = Alignment(horizontal="center", vertical="center")
    R = Alignment(horizontal="right",  vertical="center")
    FMT = 'R$ #,##0.00'; FPCT = '0.0%'

    NAVY  = "1F3864"; BLUE  = "2E75B6"; LBLUE = "BDD7EE"
    GREY  = "F2F2F2"; WHITE = "FFFFFF"

    CANAIS = ["Shopee", "Mercado Livre", "Magalu", "Amazon", "Shein", "Site Próprio", "TOTAL"]
    COR_CANAL = {
        "Shopee": "EE4D2D", "Mercado Livre": BLUE, "Magalu": "0070C0",
        "Amazon": "FF9900", "Shein": "E83E8C", "Site Próprio": "375623", "TOTAL": NAVY,
    }

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 28
    for i in range(len(CANAIS) + 1):
        ws.column_dimensions[get_column_letter(3 + i)].width = 15

    col_produto = _find_col(df.columns, "PRODUTO")
    col_total   = _find_col(df.columns, "Total (BRL)")
    col_qty     = _find_col(df.columns, "TOTAL UNIDADES")
    col_unid    = _find_col(df.columns, "Quantidade")

    def _num(c):
        return pd.to_numeric(df[c], errors="coerce").fillna(0) if c else pd.Series([0.0]*len(df))

    df["_rec"] = _num(col_total)
    df["_qty"] = _num(col_qty) if col_qty else _num(col_unid)

    if col_produto:
        por_prod = (df[df[col_produto].astype(str).str.strip() != ""]
                    .groupby(col_produto)
                    .agg(fat=("_rec","sum"), qty=("_qty","sum"))
                    .sort_values("fat", ascending=False)
                    .reset_index())
    else:
        por_prod = pd.DataFrame()

    ORDEM = [
        "Pega Areia","Tapete Lavável","Tapete Comedouro","Arranhador de Braço",
        "Arranhador de Braço Simples","Arranhador Adesivo","Arranhador de Madeira",
        "Casinha Toca","Capa para Carro","Camisa Dryfit","Casinhas / Caminhas",
        "Playground","Chuchups","Presilhas Aspirais","Piso Carpete","Sanitário",
        "Sanitario e Pega Areia","Caminha Rede","Arranhador Túnel","Degrau",
        "Caminha Comeia","Comedouro Automático","Brinquedo Interativo","Arranhadores",
        "Pato Interativo","Sanitário Automático","Comedouro Inteligente","FullSellers",
    ]
    produtos_bf = set(por_prod[col_produto].astype(str).str.strip().tolist()) if len(por_prod) else set()
    extras      = sorted([p for p in produtos_bf - set(ORDEM) if p not in ("","nan")])
    ordem_final = ORDEM + extras

    n_cols = 2 + len(CANAIS) + 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1, value="PRODUTOS  —  Faturamento por Canal de Venda")
    c.font = _f(bold=True, color="FFFFFF", size=13)
    c.fill = _bg(NAVY); c.alignment = C; ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    c = ws.cell(row=2, column=1,
                value="Shopee: dados reais  |  Demais canais: estrutura preparada para implementação futura")
    c.font = _f(size=9, color="7F6000")
    c.fill = _bg("FFF2CC"); c.alignment = C; ws.row_dimensions[2].height = 14

    c = ws.cell(row=3, column=2, value="Produto")
    c.font = _f(bold=True, color="FFFFFF", size=9)
    c.fill = _bg(NAVY); c.alignment = C; c.border = BRD
    for i, canal in enumerate(CANAIS):
        col = 3 + i
        c = ws.cell(row=3, column=col, value=canal)
        c.font = _f(bold=True, color="FFFFFF", size=9)
        c.fill = _bg(COR_CANAL.get(canal, NAVY)); c.alignment = C; c.border = BRD
    c = ws.cell(row=3, column=3+len(CANAIS), value="% Shopee s/ Total")
    c.font = _f(bold=True, color="FFFFFF", size=9)
    c.fill = _bg(COR_CANAL["Shopee"]); c.alignment = C; c.border = BRD
    ws.row_dimensions[3].height = 16

    linha = 4; linhas_fat = []; i_alt = 0
    for prod in ordem_final:
        bg = _bg(GREY if i_alt % 2 == 0 else WHITE); i_alt += 1
        if len(por_prod) > 0 and col_produto:
            rp = por_prod[por_prod[col_produto].astype(str).str.strip() == prod]
            fat_shopee = rp["fat"].sum() if len(rp) else 0
        else:
            fat_shopee = 0

        c = ws.cell(row=linha, column=2, value=prod)
        c.font = _f(size=9); c.fill = bg; c.alignment = L; c.border = BRD

        for i, canal in enumerate(CANAIS):
            col = 3 + i
            if canal == "Shopee":
                val = fat_shopee; cf = _f(size=9); cf2 = bg
            elif canal == "TOTAL":
                refs = "+".join(get_column_letter(3+j)+str(linha) for j in range(len(CANAIS)-1))
                val = f"={refs}"; cf = _f(bold=True, size=9); cf2 = _bg(LBLUE)
            else:
                val = None; cf = _f(size=9, color="9E9E9E"); cf2 = bg
            c = ws.cell(row=linha, column=col, value=val)
            c.font = cf; c.fill = cf2; c.alignment = R
            c.number_format = FMT; c.border = BRD

        col_tot = 3 + len(CANAIS) - 1
        col_pct = 3 + len(CANAIS)
        c = ws.cell(row=linha, column=col_pct,
                    value=f'=IFERROR({get_column_letter(3)}{linha}/'
                           f'{get_column_letter(col_tot)}{linha},"")')
        c.font = _f(size=9, italic=True); c.fill = bg
        c.alignment = R; c.number_format = FPCT; c.border = BRD
        ws.row_dimensions[linha].height = 14
        linhas_fat.append(linha); linha += 1

    c = ws.cell(row=linha, column=2, value="TOTAL GERAL")
    c.font = _f(bold=True, color="FFFFFF", size=10)
    c.fill = _bg(NAVY); c.alignment = L; c.border = BRD
    for i, canal in enumerate(CANAIS):
        col = 3 + i
        cl  = get_column_letter(col)
        c = ws.cell(row=linha, column=col,
                    value=f"=SUM({cl}{linhas_fat[0]}:{cl}{linhas_fat[-1]})")
        c.font = _f(bold=True, color="FFFFFF", size=10)
        c.fill = _bg(COR_CANAL.get(canal, NAVY)); c.alignment = R
        c.number_format = FMT; c.border = BRD
    ws.cell(row=linha, column=3+len(CANAIS)).fill = _bg(NAVY)
    ws.cell(row=linha, column=3+len(CANAIS)).border = BRD
    ws.row_dimensions[linha].height = 18
    ws.freeze_panes = "C4"

# =============================================================================
# ESCRITA DO XLSX FINAL
# =============================================================================

def escrever_xlsx(base_final_df, output: io.BytesIO, log, tabela_custo_fech=None):
    wb = Workbook()
    wb.remove(wb.active)

    _escrever_base_final(wb, base_final_df)
    _escrever_fechamento(wb, base_final_df, tabela_custo_fech=tabela_custo_fech)
    _escrever_painel_produtos(wb, base_final_df)

    wb.save(output)
    log.info("Arquivo gerado com sucesso")

# =============================================================================
# INTERFACE PÚBLICA
# =============================================================================

def processar(arquivos_input: list, tabela_bytes: bytes):
    """
    arquivos_input : list of (nome: str, conteudo: bytes)
    tabela_bytes   : conteúdo de TABELA_AUXILIAR.xlsx em bytes
    Retorna        : (xlsx_bytes: bytes, logs: list[str])
    """
    log = SimpleLog()

    ricapet_rows, ricapet_cols = [], []
    thapets_rows, thapets_cols = [], []

    for nome, conteudo in arquivos_input:
        empresa = identificar_empresa(nome)
        if empresa is None:
            log.warning(f"Ignorado (nome não reconhecido como relatório Shopee): {nome}")
            continue
        rows, cols = ler_arquivo(conteudo, nome, empresa, log)
        if empresa == "Ricapet":
            ricapet_rows.extend(rows)
            if not ricapet_cols:
                ricapet_cols = cols
        else:
            thapets_rows.extend(rows)
            if not thapets_cols:
                thapets_cols = cols

    if not ricapet_rows and not thapets_rows:
        raise ValueError(
            "Nenhum dado foi lido. Verifique se os arquivos são relatórios Shopee "
            "(o nome deve começar com 'order.all.'). "
            f"Ricapet: contém '{CHAVE_RICAPET}' no nome. Thapets: não contém."
        )

    if not ricapet_cols: ricapet_cols = thapets_cols
    if not thapets_cols: thapets_cols = ricapet_cols

    rows, all_cols = unificar(ricapet_rows, ricapet_cols, thapets_rows, thapets_cols, log)
    rows = distribuir_valores(rows, all_cols, log)
    rows = calcular_total_liquido(rows, all_cols, log)

    log.info("Carregando TABELA_AUXILIAR...")
    tabela_produtos, tabela_status, tabela_custo_fech, tabela_custo_unit = carregar_auxiliares(tabela_bytes, log)

    log.info("Criando Base Final...")
    base_final_df = criar_base_final(rows, all_cols, log)

    log.info("Preenchendo colunas calculadas...")
    base_final_df = preencher_colunas(
        base_final_df, tabela_produtos, tabela_status, log,
        tabela_custo_unit=tabela_custo_unit
    )

    log.info("Gerando arquivo Excel...")
    output = io.BytesIO()
    escrever_xlsx(base_final_df, output, log, tabela_custo_fech=tabela_custo_fech)
    output.seek(0)

    log.info(f"Concluído: {len(base_final_df)} linhas na Base Final")
    return output.read(), log.msgs
