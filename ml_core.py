# ml_core.py
# ==========
# Versão do consolidar_mercadolivre.py adaptada para rodar em memória (Streamlit).
# Não lê/grava arquivos em disco — recebe bytes e retorna bytes.
#
# Interface pública:
#   processar(arquivos, tabela_bytes) -> (xlsx_bytes: bytes, logs: list[str])
#
# Dependências: openpyxl, pandas

import io
import math
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# CONSTANTES
# =============================================================================

CHAVE_RICAPET  = "736787693"
CHAVE_THAPETS  = "1139210125"

COR_MAE        = "FFB7B7B7"
COR_FILHO      = "FFF3F3F3"
COR_NENHUMA    = "00000000"

TOLERANCIA     = 0.02

COL_RECEITA_PROD  = "Receita por produtos (BRL)"
COL_RECEITA_ENVIO = "Receita por envio (BRL)"
COL_TARIFA        = "Tarifa de venda e impostos (BRL)"
COL_TARIFAS_ENVIO = "Tarifas de envio (BRL)"
COL_CANCELAMENTOS = "Cancelamentos e reembolsos (BRL)"
COL_TOTAL         = "Total (BRL)"
COL_UNIDADES      = "Unidades"
COL_PRECO_UNIT    = "Preço unitário de venda do anúncio (BRL)"
COL_ORDER         = "N.º de venda"
COL_LOJA_OFICIAL  = "Loja oficial"
COL_DESCONTOS     = "Descontos e bônus"

COLUNAS_TOTAL = [
    "Receita por produtos (BRL)",
    "Receita por acréscimo no preço (pago pelo comprador)",
    "Taxa de parcelamento equivalente ao acréscimo",
    "Tarifa de venda e impostos (BRL)",
    "Receita por envio (BRL)",
    "Tarifas de envio (BRL)",
    "Custo de envio com base nas medidas e peso declarados",
    "Custo por diferenças nas medidas e no peso do pacote",
    "Descontos e bônus",
    "Cancelamentos e reembolsos (BRL)",
]

COLUNAS_VALIDAR = [
    COL_RECEITA_PROD, COL_RECEITA_ENVIO, COL_TARIFA, COL_TARIFAS_ENVIO,
    COL_CANCELAMENTOS, COL_DESCONTOS, COL_TOTAL, COL_UNIDADES,
]

FILHO_OWNS = {
    COL_ORDER, "SKU", "# de anúncio", "Canal de venda", "Título do anúncio",
    "Variação", COL_PRECO_UNIT, "Tipo de anúncio", COL_UNIDADES,
    "Estado", "Descrição do status", "Reclamação aberta",
    "Reclamação encerrada", "Em mediação", "Pedido de compra",
    "Venda por publicidade", "Pacote de diversos produtos",
    "Pertence a um kit",
}

# =============================================================================
# LOG SIMPLES
# =============================================================================

class SimpleLog:
    def __init__(self):
        self.msgs = []

    def info(self, msg):
        self.msgs.append(f"ℹ️  {msg}")

    def warning(self, msg):
        self.msgs.append(f"⚠️  {msg}")

    def error(self, msg):
        self.msgs.append(f"❌  {msg}")

# =============================================================================
# UTILITÁRIOS
# =============================================================================

def sf(v, default=0.0):
    if v is None:
        return default
    if isinstance(v, float) and math.isnan(v):
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default

def vazio(v):
    if v is None:
        return True
    if isinstance(v, float) and math.isnan(v):
        return True
    return str(v).strip() in ("", " ", "nan")

def arr(v):
    return round(v, 2)

# =============================================================================
# LEITURA DOS RELATÓRIOS
# =============================================================================

def identificar_empresa(nome: str):
    if CHAVE_RICAPET in nome:
        return "Ricapet"
    if CHAVE_THAPETS in nome:
        return "Thapets"
    return None

def ler_arquivo(conteudo: bytes, nome: str, empresa: str, log):
    log.info(f"Lendo {nome}  [{empresa}]")
    wb = load_workbook(io.BytesIO(conteudo))
    ws = wb.active

    header_row = None
    for r in range(1, 15):
        if ws.cell(row=r, column=1).value == COL_ORDER:
            header_row = r
            break
    if header_row is None:
        raise ValueError(f"Cabeçalho '{COL_ORDER}' não encontrado em {nome}")

    raw_headers = []
    seen = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=header_row, column=c).value
        if not h:
            h = f"_col_{c}"
        h = str(h).strip()
        seen[h] = seen.get(h, 0) + 1
        if seen[h] > 1:
            h = f"{h}_{seen[h]}"
        raw_headers.append(h)

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        v1 = ws.cell(row=r, column=1).value
        if v1 is None:
            continue
        rgb = ws.cell(row=r, column=1).fill.fgColor.rgb
        row = {"_cor": rgb, "_empresa": empresa}
        for c, h in enumerate(raw_headers, 1):
            row[h] = ws.cell(row=r, column=c).value
        rows.append(row)

    log.info(f"  → {len(rows)} linhas  |  {len(raw_headers)} colunas")
    return rows, raw_headers

# =============================================================================
# UNIFICAÇÃO
# =============================================================================

def empresa_por_loja(row, fallback):
    loja = row.get(COL_LOJA_OFICIAL)
    if loja is None or str(loja).strip() == "":
        return fallback
    return "Ricapet" if str(loja).strip().lower() == "ricapet" else "Thapets"

def unificar(ricapet_rows, ricapet_cols, thapets_rows, thapets_cols, log):
    extra    = [c for c in thapets_cols if c not in ricapet_cols]
    all_cols = ["Empresa"] + ricapet_cols + extra

    def norm(rows):
        out = []
        for r in rows:
            empresa = empresa_por_loja(r, fallback=r["_empresa"])
            row = {"Empresa": empresa, "_cor": r["_cor"], "_empresa": empresa}
            for c in ricapet_cols + extra:
                row[c] = r.get(c)
            out.append(row)
        return out

    unified    = norm(ricapet_rows) + norm(thapets_rows)
    ricapet_ct = sum(1 for r in unified if r["Empresa"] == "Ricapet")
    thapets_ct = sum(1 for r in unified if r["Empresa"] == "Thapets")
    log.info(f"Unificado: {len(unified)} linhas  |  {len(all_cols)} colunas")
    log.info(f"  Empresa → Ricapet: {ricapet_ct}  |  Thapets: {thapets_ct}")
    return unified, all_cols

# =============================================================================
# PROCESSAMENTO DE GRUPOS MÃE / FILHOS
# =============================================================================

def construir_grupos(rows):
    groups = []
    i = 0
    while i < len(rows):
        if rows[i]["_cor"] == COR_MAE:
            filhos = []
            j = i + 1
            while j < len(rows) and rows[j]["_cor"] == COR_FILHO:
                filhos.append(j)
                j += 1
            groups.append((i, filhos))
            i = j
        else:
            i += 1
    return groups

def processar_grupos(rows, groups, log):
    for r in rows:
        r["Tem"] = None

    for mae_idx, filhos_idx in groups:
        mae = rows[mae_idx]
        mae["Tem"] = 1
        if not filhos_idx:
            continue
        n = len(filhos_idx)
        for fi in filhos_idx:
            rows[fi]["Tem"] = 2

        for fi in filhos_idx:
            preco = sf(rows[fi].get(COL_PRECO_UNIT), default=None)
            unid  = sf(rows[fi].get(COL_UNIDADES),   default=None)
            if preco is not None and unid is not None and preco != 0:
                rows[fi][COL_RECEITA_PROD] = arr(preco * unid)

        for fi in filhos_idx:
            for col, val in mae.items():
                if col in ("_cor", "_empresa", "Tem", "Empresa") or col in FILHO_OWNS:
                    continue
                if vazio(rows[fi].get(col)) and not vazio(val):
                    rows[fi][col] = val

        if vazio(mae.get(COL_LOJA_OFICIAL)):
            for fi in filhos_idx:
                loja_filho = rows[fi].get(COL_LOJA_OFICIAL)
                if not vazio(loja_filho):
                    mae[COL_LOJA_OFICIAL] = loja_filho
                    break
        mae["Empresa"]  = empresa_por_loja(mae, fallback=mae["_empresa"])
        mae["_empresa"] = mae["Empresa"]
        for fi in filhos_idx:
            rows[fi]["Empresa"]  = empresa_por_loja(rows[fi], fallback=rows[fi]["_empresa"])
            rows[fi]["_empresa"] = rows[fi]["Empresa"]

        total_unid = sum(sf(rows[fi].get(COL_UNIDADES)) for fi in filhos_idx)
        if total_unid > 0:
            mae[COL_UNIDADES] = int(total_unid) if total_unid == int(total_unid) else total_unid

        COLUNAS_PROPORCIONAL_TODOS = [
            "Receita por acréscimo no preço (pago pelo comprador)",
            "Taxa de parcelamento equivalente ao acréscimo",
            COL_TARIFA, COL_RECEITA_ENVIO, COL_TARIFAS_ENVIO,
            "Custo de envio com base nas medidas e peso declarados",
            "Custo por diferenças nas medidas e no peso do pacote",
            COL_DESCONTOS, COL_CANCELAMENTOS,
        ]

        mae_receita = sf(mae.get(COL_RECEITA_PROD))
        if mae_receita != 0:
            pcts = [sf(rows[fi].get(COL_RECEITA_PROD)) / mae_receita for fi in filhos_idx]
            for col in COLUNAS_PROPORCIONAL_TODOS:
                mae_val = sf(mae.get(col), default=None)
                if mae_val is None:
                    continue
                valores = [arr(mae_val * p) for p in pcts]
                diff = arr(mae_val - arr(sum(valores)))
                valores[-1] = arr(valores[-1] + diff)
                for fi, v in zip(filhos_idx, valores):
                    rows[fi][col] = v

        for fi in filhos_idx:
            rows[fi][COL_TOTAL] = arr(sum(sf(rows[fi].get(c)) for c in COLUNAS_TOTAL))

    mae_ct   = sum(1 for r in rows if r.get("Tem") == 1)
    filho_ct = sum(1 for r in rows if r.get("Tem") == 2)
    log.info(f"Processamento: {mae_ct} mães  |  {filho_ct} filhos  |  {len(groups)} grupos")
    return rows

# =============================================================================
# VALIDAÇÃO
# =============================================================================

def gerar_erros(rows, groups, log):
    erros = []
    for mae_idx, filhos_idx in groups:
        if not filhos_idx:
            continue
        mae     = rows[mae_idx]
        order   = mae.get(COL_ORDER, "?")
        empresa = mae.get("_empresa", "?")
        f_linhas = ", ".join(str(fi + 2) for fi in filhos_idx)

        for col in COLUNAS_VALIDAR:
            mae_val = sf(mae.get(col), default=None)
            if mae_val is None:
                continue
            soma = arr(sum(sf(rows[fi].get(col)) for fi in filhos_idx))
            diff = arr(mae_val - soma)
            if abs(diff) > TOLERANCIA:
                erros.append({
                    "Número do Pedido":    order,
                    "Linha Mãe":           mae_idx + 2,
                    "Linhas Filhos":       f_linhas,
                    "Empresa":             empresa,
                    "Campo Analisado":     col,
                    "Valor Pedido Mãe":    arr(mae_val),
                    "Soma Pedidos Filhos": soma,
                    "Diferença":           diff,
                })
    log.info(f"Validação: {len(erros)} divergência(s)")
    return erros

# =============================================================================
# CARREGAMENTO DA TABELA AUXILIAR
# =============================================================================

def carregar_auxiliares(tabela_bytes: bytes, log):
    xl = pd.ExcelFile(io.BytesIO(tabela_bytes))

    def norm_sheet(name):
        import unicodedata
        n = unicodedata.normalize("NFKD", str(name))
        n = "".join(c for c in n if not unicodedata.combining(c))
        return n.strip().upper()

    sheet_map = {norm_sheet(s): s for s in xl.sheet_names}
    log.info(f"  Abas encontradas em TABELA_AUXILIAR: {xl.sheet_names}")

    chave_tp = norm_sheet("TABELA_PRODUTOS")
    if chave_tp not in sheet_map:
        raise ValueError(
            f"Aba 'TABELA_PRODUTOS' não encontrada em TABELA_AUXILIAR.xlsx.\n"
            f"Abas disponíveis: {xl.sheet_names}"
        )
    tp = xl.parse(sheet_map[chave_tp], dtype=str).fillna("")
    tp.columns = [str(c).strip() for c in tp.columns]
    col_map_tp = {str(c).strip().upper(): c for c in tp.columns}
    log.info(f"  Colunas em TABELA_PRODUTOS: {list(tp.columns)}")
    for col in ["SKU", "TÍTULO", "PRODUTO", "COR", "TAMANHO", "UNIDADE"]:
        if col not in col_map_tp:
            raise ValueError(
                f"Coluna '{col}' não encontrada em TABELA_PRODUTOS.\n"
                f"Colunas disponíveis: {list(tp.columns)}"
            )
    rename_tp = {}
    for canon, upper in [("SKU", "SKU"), ("Título", "TÍTULO"),
                         ("PRODUTO", "PRODUTO"), ("COR", "COR"),
                         ("TAMANHO", "TAMANHO"), ("UNIDADE", "UNIDADE")]:
        real = col_map_tp[upper]
        if real != canon:
            rename_tp[real] = canon
    if rename_tp:
        tp = tp.rename(columns=rename_tp)

    chave_st = norm_sheet("STATUS")
    if chave_st not in sheet_map:
        raise ValueError(
            f"Aba 'STATUS' não encontrada em TABELA_AUXILIAR.xlsx.\n"
            f"Abas disponíveis: {xl.sheet_names}"
        )
    st = xl.parse(sheet_map[chave_st], dtype=str).fillna("")
    st.columns = [str(c).strip() for c in st.columns]
    col_map_st = {str(c).strip().upper(): c for c in st.columns}
    log.info(f"  Colunas em STATUS: {list(st.columns)}")

    col_status = col_map_st.get("STATUS")
    col_manter = col_map_st.get("MANTER")
    if col_status is None or col_manter is None:
        raise ValueError(
            f"Aba STATUS precisa das colunas 'Status' e 'Manter'.\n"
            f"Colunas encontradas: {list(st.columns)}"
        )
    rename_st = {}
    if col_status != "Status":
        rename_st[col_status] = "Status"
    if col_manter != "Manter":
        rename_st[col_manter] = "Manter"
    if rename_st:
        st = st.rename(columns=rename_st)

    def _parse_custo_aba(xl_obj, nome_aba, sheet_m, log_prefix):
        chave = norm_sheet(nome_aba)
        if chave not in sheet_m:
            return pd.DataFrame()
        raw = xl_obj.parse(sheet_m[chave], header=None)
        hdr = None
        for i_r, row_r in raw.iterrows():
            if any(str(v).strip().upper() == "PRODUTO" for v in row_r if pd.notna(v)):
                hdr = i_r; break
        if hdr is None:
            raw2 = xl_obj.parse(sheet_m[chave])
            raw2.columns = [str(c).strip() for c in raw2.columns]
            cm = {str(c).strip().upper(): c for c in raw2.columns}
            if "PRODUTO" not in cm:
                return pd.DataFrame()
            raw = raw2; hdr = -1
        if hdr >= 0:
            raw.columns = [str(v).strip() for v in raw.iloc[hdr]]
            raw = raw.iloc[hdr + 1:].reset_index(drop=True)
        cm = {str(c).strip().upper(): c for c in raw.columns}
        c_prod  = cm.get("PRODUTO")
        c_tam   = cm.get("TAMANHO")
        c_custo = next((v for k, v in cm.items() if "CUSTO" in k and "UNIT" in k),
                       next((v for k, v in cm.items() if "CUSTO" in k), None))
        if not c_prod or not c_custo:
            return pd.DataFrame()
        ren = {}
        if c_prod  != "PRODUTO":         ren[c_prod]  = "PRODUTO"
        if c_tam and c_tam != "TAMANHO": ren[c_tam]   = "TAMANHO"
        if c_custo != "CUSTO_UNITARIO":  ren[c_custo] = "CUSTO_UNITARIO"
        if ren: raw = raw.rename(columns=ren)
        if "TAMANHO" not in raw.columns: raw["TAMANHO"] = ""
        raw["PRODUTO"]        = raw["PRODUTO"].astype(str).str.strip()
        raw["TAMANHO"]        = raw["TAMANHO"].astype(str).str.strip().replace("nan", "")
        raw["CUSTO_UNITARIO"] = pd.to_numeric(raw["CUSTO_UNITARIO"], errors="coerce")
        raw = raw[~raw["PRODUTO"].str.upper().str.contains("TOTAL", na=False)]
        raw = raw[raw["PRODUTO"].str.strip() != ""]
        raw = raw.dropna(subset=["CUSTO_UNITARIO"])
        result = raw[["PRODUTO", "TAMANHO", "CUSTO_UNITARIO"]].reset_index(drop=True)
        log.info(f"  {log_prefix}: {len(result)} registros carregados")
        return result

    cu_precif    = _parse_custo_aba(xl, "CUSTO_UNITARIO", sheet_map, "CUSTO_UNITARIO")
    cu_planilha2 = _parse_custo_aba(xl, "Planilha2",      sheet_map, "Planilha2 (custos reais)")

    if len(cu_precif) > 0 and len(cu_planilha2) > 0:
        cu_base = cu_precif.set_index(["PRODUTO", "TAMANHO"])
        cu_over = cu_planilha2.set_index(["PRODUTO", "TAMANHO"])
        cu_base.update(cu_over)
        novos = cu_over[~cu_over.index.isin(cu_base.index)]
        cu = pd.concat([cu_base, novos]).reset_index()
    elif len(cu_planilha2) > 0:
        cu = cu_planilha2
    elif len(cu_precif) > 0:
        cu = cu_precif
    else:
        cu = pd.DataFrame(columns=["PRODUTO", "TAMANHO", "CUSTO_UNITARIO"])
        log.warning("  Nenhuma fonte de custo encontrada em TABELA_AUXILIAR.xlsx.")

    chave_custo_aba = norm_sheet("CUSTO")
    tabela_custo_fech = pd.DataFrame()
    if chave_custo_aba in sheet_map:
        tc = xl.parse(sheet_map[chave_custo_aba]).fillna("")
        tc.columns = [str(c).strip() for c in tc.columns]
        cm = {str(c).strip().upper(): c for c in tc.columns}
        c_prod = cm.get("PRODUTO"); c_cor = cm.get("COR")
        c_tam  = cm.get("TAMANHO"); c_cst = cm.get("CUSTO")
        if c_prod and c_cst:
            ren = {}
            if c_prod != "PRODUTO":              ren[c_prod] = "PRODUTO"
            if c_cor  and c_cor != "COR":        ren[c_cor]  = "COR"
            if c_tam  and c_tam != "TAMANHO":    ren[c_tam]  = "TAMANHO"
            if c_cst  != "CUSTO":                ren[c_cst]  = "CUSTO"
            if ren: tc = tc.rename(columns=ren)
            for col in ["COR", "TAMANHO"]:
                if col not in tc.columns: tc[col] = ""
            tc["PRODUTO"]  = tc["PRODUTO"].astype(str).str.strip()
            tc["COR"]      = tc["COR"].astype(str).str.strip()
            tc["TAMANHO"]  = tc["TAMANHO"].astype(str).str.strip()
            tc["CUSTO"]    = pd.to_numeric(tc["CUSTO"], errors="coerce")
            tc = tc[tc["PRODUTO"].str.strip() != ""].dropna(subset=["CUSTO"])
            tabela_custo_fech = tc[["PRODUTO","COR","TAMANHO","CUSTO"]].reset_index(drop=True)
            log.info(f"  Aba CUSTO carregada: {len(tabela_custo_fech)} registros")
        else:
            log.warning("  Aba CUSTO: colunas PRODUTO ou CUSTO não localizadas")
    else:
        log.warning("  Aba CUSTO não encontrada em TABELA_AUXILIAR.xlsx")

    log.info(
        f"TABELA_AUXILIAR carregada: {len(tp)} produtos  |  "
        f"{len(st)} status  |  {len(cu)} custos Base  |  "
        f"{len(tabela_custo_fech)} custos Fechamento"
    )
    return tp, st, cu, tabela_custo_fech

# =============================================================================
# CRIAÇÃO DA BASE FINAL
# =============================================================================

def criar_base_final(rows, all_cols, log):
    data_cols = [c for c in all_cols if c != "Empresa"]
    base_cols = ["Tem", "Empresa"] + data_cols

    registros = []
    for row in rows:
        if row.get("Tem") == 1:
            continue
        reg = {}
        for col in base_cols:
            if col == "Tem":
                reg[col] = row.get("Tem")
            elif col == "Empresa":
                reg[col] = row.get("Empresa", row.get("_empresa", ""))
            else:
                val = row.get(col)
                reg[col] = None if (isinstance(val, float) and math.isnan(val)) else val
        registros.append(reg)

    df = pd.DataFrame(registros, columns=base_cols)

    col_order = next((c for c in df.columns if str(c).strip() == "N.º de venda"), None)
    col_total  = next((c for c in df.columns if str(c).strip() == "Total (BRL)"), None)
    col_sku    = next((c for c in df.columns if str(c).strip() == "SKU"), None)

    if col_order and col_total and col_sku:
        sem_tem = df[df["Tem"].isna()]
        duplicados = sem_tem[sem_tem.duplicated(subset=[col_order], keep=False)]
        pedidos_dup = duplicados[col_order].unique()

        if len(pedidos_dup) > 0:
            log.warning(f"  {len(pedidos_dup)} pedido(s) com linhas duplicadas sem mãe/filho — consolidando")
            idx_remover = []
            for ped in pedidos_dup:
                grupo = df[df[col_order].astype(str).str.strip() == str(ped).strip()]
                com_sku  = grupo[grupo[col_sku].astype(str).str.strip() != ""]
                sem_sku  = grupo[grupo[col_sku].astype(str).str.strip() == ""]
                if len(com_sku) == 0 or len(sem_sku) == 0:
                    continue
                total_extra = pd.to_numeric(sem_sku[col_total].fillna(0), errors="coerce").sum()
                for idx in com_sku.index:
                    atual = pd.to_numeric(df.at[idx, col_total], errors="coerce")
                    if math.isnan(atual) or atual == 0:
                        df.at[idx, col_total] = round(total_extra, 2)
                    else:
                        df.at[idx, col_total] = round(float(atual) + total_extra, 2)
                idx_remover.extend(sem_sku.index.tolist())

            if idx_remover:
                df = df.drop(index=idx_remover).reset_index(drop=True)
                log.warning(f"  {len(idx_remover)} linha(s) duplicada(s) removida(s)")

    log.info(f"Base Final: {len(df)} linhas (pedidos mãe removidos)")
    return df

# =============================================================================
# PREENCHIMENTO DAS NOVAS COLUNAS
# =============================================================================

def preencher_colunas(df, tabela_produtos, tabela_status, log, tabela_custo=None):
    for col in ["PRODUTO", "COR", "TAMANHO", "UNIDADE", "MANTER"]:
        df[col] = ""
    df["Custo"]  = None
    df["Alerta"] = ""

    campos_str = ["PRODUTO", "COR", "TAMANHO", "UNIDADE"]

    _col_custo_tp = next(
        (c for c in tabela_produtos.columns if str(c).strip().upper() == "CUSTO"), None
    )
    if _col_custo_tp is None:
        log.warning("Coluna 'CUSTO' não encontrada em TABELA_PRODUTOS — coluna Custo ficará vazia.")

    lookup_sku    = {}
    lookup_titulo = {}
    for _, row in tabela_produtos.iterrows():
        sku    = str(row.get("SKU",    "")).strip()
        titulo = str(row.get("Título", "")).strip()
        dados_str  = {c: str(row.get(c, "")).strip() for c in campos_str}
        _raw_custo = row.get(_col_custo_tp, "") if _col_custo_tp else ""
        _raw_str   = str(_raw_custo).strip()
        if _raw_str in ("", "nan", "NaN", "None", "-"):
            _custo_val = None
        else:
            try:
                _custo_val = float(_raw_str)
            except (ValueError, TypeError):
                _custo_val = None
        dados_str["CUSTO"] = _custo_val
        if sku and sku.lower() not in lookup_sku:
            lookup_sku[sku.lower()] = dados_str
        if titulo and titulo.lower() not in lookup_titulo:
            lookup_titulo[titulo.lower()] = dados_str

    lookup_status = {
        str(row["Status"]).strip().lower(): str(row["Manter"]).strip()
        for _, row in tabela_status.iterrows()
        if str(row.get("Status", "")).strip()
    }

    col_sku    = next((c for c in df.columns if str(c).strip().lower() == "sku"), None)
    col_titulo = next((c for c in df.columns
                       if str(c).strip().lower() in
                       ("título do anúncio", "titulo do anuncio", "título", "titulo")), None)
    col_estado = next((c for c in df.columns if str(c).strip().lower() == "estado"), None)

    if col_sku is None:
        raise ValueError("Coluna 'SKU' não encontrada na Base Final.")

    acertos_sku = acertos_titulo = acertos_status = 0

    for idx in df.index:
        sku    = str(df.at[idx, col_sku]).strip()
        titulo = str(df.at[idx, col_titulo]).strip() if col_titulo else ""
        estado = str(df.at[idx, col_estado]).strip() if col_estado else ""

        dados = lookup_sku.get(sku.lower())
        if dados:
            for c in campos_str:
                df.at[idx, c] = dados[c]
            if dados["CUSTO"] is not None:
                df.at[idx, "Custo"] = dados["CUSTO"]
            acertos_sku += 1
        elif titulo and col_titulo:
            dados = lookup_titulo.get(titulo.lower())
            if dados:
                for c in campos_str:
                    df.at[idx, c] = dados[c]
                if dados["CUSTO"] is not None:
                    df.at[idx, "Custo"] = dados["CUSTO"]
                acertos_titulo += 1

        if estado and col_estado:
            manter = lookup_status.get(estado.lower(), "")
            df.at[idx, "MANTER"] = manter
            if manter:
                acertos_status += 1

    import unicodedata as _ud2
    def _norm2(s):
        s = str(s).strip()
        s = _ud2.normalize("NFKD", s)
        s = "".join(ch for ch in s if not _ud2.combining(ch))
        return " ".join(s.lower().split())

    if tabela_custo is not None and not tabela_custo.empty:
        lk_prod_tam = {}
        lk_prod_only = {}
        for _, rc in tabela_custo.iterrows():
            pk = _norm2(rc.get("PRODUTO",""))
            tk = _norm2(rc.get("TAMANHO",""))
            cv = rc.get("CUSTO_UNITARIO")
            if pk and cv is not None and not (isinstance(cv, float) and cv != cv):
                lk_prod_tam[(pk, tk)] = float(cv)
                if tk in ("", "-", "—"):
                    lk_prod_only[pk] = float(cv)

        col_prod_bf = next((c for c in df.columns if str(c).strip().upper() == "PRODUTO"), None)
        col_tam_bf  = next((c for c in df.columns if str(c).strip().upper() == "TAMANHO"), None)

        if col_prod_bf:
            acertos_pt = 0
            for idx in df.index:
                if df.at[idx, "Custo"] is not None:
                    continue
                pk = _norm2(df.at[idx, col_prod_bf])
                tk = _norm2(df.at[idx, col_tam_bf]) if col_tam_bf else ""
                custo_v = (lk_prod_tam.get((pk, tk))
                           or lk_prod_tam.get((pk, ""))
                           or lk_prod_tam.get((pk, "-"))
                           or lk_prod_only.get(pk))
                if custo_v is not None:
                    df.at[idx, "Custo"] = custo_v
                    acertos_pt += 1
            log.info(f"  Cruzamento PRODUTO+TAMANHO (tabela_custo): {acertos_pt} registros preenchidos")

    col_total_bf = next((c for c in df.columns if str(c).strip() == "Total (BRL)"), None)
    if col_total_bf:
        total_num = pd.to_numeric(df[col_total_bf], errors="coerce")
        df["Alerta"] = total_num.apply(
            lambda v: "⚠" if (v is not None and not (isinstance(v, float) and v != v) and v <= 0) else ""
        )
        n_alertas = int((df["Alerta"] == "⚠").sum())
        if n_alertas:
            log.warning(f"  {n_alertas} linha(s) com Total (BRL) <= 0 marcadas na coluna Alerta")

    acertos_custo = int(df["Custo"].notna().sum())
    log.info(f"Cruzamento SKU: {acertos_sku}  |  Título: {acertos_titulo}  |  "
             f"MANTER: {acertos_status}  |  Custo preenchido: {acertos_custo}")

    col_unidades = next((c for c in df.columns if str(c).strip().lower() == "unidades"), None)
    col_unidade  = next((c for c in df.columns if str(c).strip().upper() == "UNIDADE"), None)
    if col_unidades and col_unidade:
        unid_num = pd.to_numeric(df[col_unidades], errors="coerce").fillna(0)
        comp_num = pd.to_numeric(df[col_unidade],  errors="coerce").fillna(1)
        df["TOTAL UNIDADES"] = (unid_num * comp_num).round(0).astype(int)
    else:
        df["TOTAL UNIDADES"] = None
        log.warning("TOTAL UNIDADES não calculado: colunas 'Unidades' ou 'UNIDADE' não encontradas.")

    return df

# =============================================================================
# ESCRITA DO XLSX
# =============================================================================

def _hdr(ws, row, n_cols, bg="1F4E79", fg="FFFFFF"):
    fill  = PatternFill("solid", start_color=bg)
    font  = Font(name="Calibri", bold=True, color=fg, size=11)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin  = Side(style="thin", color="BFBFBF")
    brd   = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font; cell.fill = fill
        cell.alignment = align; cell.border = brd
    ws.row_dimensions[row].height = 30

def escrever_xlsx(rows, all_cols, erros, base_final_df, output: io.BytesIO, log, tabela_custo=None, tabela_custo_fech=None):
    wb = Workbook()
    wb.remove(wb.active)

    if erros:
        ws_v = wb.create_sheet("Validação")
        val_headers = [
            "Número do Pedido", "Linha Mãe", "Linhas Filhos",
            "Empresa", "Campo Analisado",
            "Valor Pedido Mãe", "Soma Pedidos Filhos", "Diferença",
        ]
        NV = len(val_headers)
        for ci, h in enumerate(val_headers, 1):
            ws_v.cell(row=1, column=ci, value=h)
        _hdr(ws_v, 1, NV)

        thin  = Side(style="thin", color="BFBFBF")
        BRD   = Border(left=thin, right=thin, top=thin, bottom=thin)
        F_ERR  = PatternFill("solid", start_color="FFCCCC")
        F_DIFF = PatternFill("solid", start_color="FFF2CC")
        BOLD   = Font(name="Calibri", size=11, bold=True)
        NORM_V = Font(name="Calibri", size=11)

        for ri, e in enumerate(erros, 2):
            vals = [
                e["Número do Pedido"], e["Linha Mãe"], e["Linhas Filhos"],
                e["Empresa"], e["Campo Analisado"],
                e["Valor Pedido Mãe"], e["Soma Pedidos Filhos"], e["Diferença"],
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws_v.cell(row=ri, column=ci, value=v)
                cell.font   = BOLD if ci in (5, 8) else NORM_V
                cell.fill   = F_DIFF if ci == 8 else F_ERR
                cell.border = BRD
                if ci in (6, 7, 8):
                    cell.alignment     = Alignment(horizontal="right",  vertical="center")
                    cell.number_format = "#,##0.00"
                elif ci == 2:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left",   vertical="center")

        for ci, w in zip(range(1, NV + 1), [28, 10, 22, 12, 36, 18, 20, 14]):
            ws_v.column_dimensions[get_column_letter(ci)].width = w
        ws_v.freeze_panes = "A2"
        log.warning(f"Aba Validação criada: {len(erros)} divergência(s) encontrada(s)")
    else:
        log.info("Sem divergências — aba Validação não gerada")

    ws_bf = wb.create_sheet("Base Final")

    COLS_FRONT = ["Alerta", "Empresa", "PRODUTO", "COR", "TAMANHO",
                  "UNIDADE", "MANTER", "Custo", "TOTAL UNIDADES"]
    todas = list(base_final_df.columns)
    colunas_bf = (
        [c for c in COLS_FRONT if c in todas]
        + [c for c in todas if c not in COLS_FRONT]
    )
    df_bf = base_final_df[colunas_bf].copy()

    _thin   = Side(style="thin",   color="E0E0E0")
    _BRD    = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _LA     = Alignment(horizontal="left",   vertical="center")
    _CA     = Alignment(horizontal="center", vertical="center")
    _RA     = Alignment(horizontal="right",  vertical="center")
    def _FNT(bold=False, color="1A1A1A", size=9):
        return Font(name="Calibri", bold=bold, color=color, size=size)

    _HDR_ORIG = "404040"
    _HDR_CALC = "1F4E79"
    _HDR_ALRT = "C00000"
    _ROW_ALRT = PatternFill("solid", start_color="FFE5E5")
    _ROW_NORM = None

    COLS_CALC = {"PRODUTO","COR","TAMANHO","UNIDADE","MANTER","Custo","TOTAL UNIDADES"}
    N_BF = len(colunas_bf)
    ws_bf.row_dimensions[1].height = 20

    for ci, h in enumerate(colunas_bf, 1):
        cell = ws_bf.cell(row=1, column=ci, value=h)
        if h == "Alerta":
            bg_hdr = _HDR_ALRT
        elif h in COLS_CALC or h == "Empresa":
            bg_hdr = _HDR_CALC
        else:
            bg_hdr = _HDR_ORIG
        cell.fill      = PatternFill("solid", start_color=bg_hdr)
        cell.font      = _FNT(bold=True, color="FFFFFF", size=9)
        cell.alignment = _CA
        cell.border    = _BRD

    col_alerta_idx = (colunas_bf.index("Alerta") if "Alerta" in colunas_bf else None)

    for ri, row_data in enumerate(df_bf.itertuples(index=False), start=2):
        row_list  = list(row_data)
        tem_alerta = (str(row_list[col_alerta_idx]).strip() == "⚠"
                      if col_alerta_idx is not None else False)
        row_fill  = _ROW_ALRT if tem_alerta else _ROW_NORM

        for ci, (col, valor) in enumerate(zip(colunas_bf, row_list), start=1):
            val  = None if (isinstance(valor, float) and math.isnan(valor)) else valor
            cell = ws_bf.cell(row=ri, column=ci, value=val)
            cell.font   = _FNT(
                bold  = (col == "Alerta" and tem_alerta),
                color = "C00000" if (col == "Alerta" and tem_alerta) else "1A1A1A",
            )
            if row_fill:
                cell.fill = row_fill
            cell.border = _BRD

            if col == "Alerta":
                cell.alignment = _CA
            elif isinstance(val, (int, float)):
                cell.alignment     = _RA
                cell.number_format = "#,##0.00"
            else:
                cell.alignment = _LA

        ws_bf.row_dimensions[ri].height = 14

    n_front = len([c for c in COLS_FRONT if c in todas])
    ws_bf.freeze_panes  = f"{get_column_letter(n_front + 1)}2"
    ws_bf.auto_filter.ref = ws_bf.dimensions

    LARG_BF = {
        "Alerta": 6,
        "Empresa": 10, "PRODUTO": 24, "COR": 13, "TAMANHO": 11,
        "UNIDADE": 9, "MANTER": 9, "Custo": 11, "TOTAL UNIDADES": 11,
        "Tem": 5, COL_ORDER: 20, "Data da venda": 18,
        "Estado": 22, "Descrição do status": 36, "Comprador": 22,
        COL_TOTAL: 12, "SKU": 32, "Título do anúncio": 38,
        "Canal de venda": 14, COL_UNIDADES: 9,
        COL_RECEITA_PROD: 18, COL_RECEITA_ENVIO: 18,
        COL_TARIFA: 24, COL_TARIFAS_ENVIO: 18, COL_CANCELAMENTOS: 22,
    }
    for ci, col in enumerate(colunas_bf, 1):
        ws_bf.column_dimensions[get_column_letter(ci)].width = LARG_BF.get(col, 14)

    _escrever_paineis(wb, base_final_df, tabela_custo=tabela_custo, tabela_custo_fech=tabela_custo_fech)
    wb.save(output)
    log.info("Arquivo gerado com sucesso")

# =============================================================================
# ABA FECHAMENTO
# =============================================================================

def _escrever_fechamento(wb, base_final_df, tabela_custo_fech=None):
    import unicodedata as _ud

    ws = wb.create_sheet("Fechamento")
    df = base_final_df.copy()

    def _col(df, *names):
        for n in names:
            m = next((x for x in df.columns if str(x).strip().lower() == n.lower()), None)
            if m: return m
        return None

    col_p  = _col(df,"PRODUTO"); col_c=_col(df,"COR"); col_t=_col(df,"TAMANHO")
    col_r  = _col(df,"Total (BRL)")
    col_qu = _col(df,"TOTAL UNIDADES") or _col(df,"Unidades")

    def _n(c): return pd.to_numeric(df[c],errors="coerce").fillna(0) if c else pd.Series([0.0]*len(df))
    df["_rec"]=_n(col_r); df["_qty"]=_n(col_qu)

    _ca=next((c for c in df.columns if str(c).strip()=="Alerta"),None)
    if _ca: df=df[df[_ca].astype(str).str.strip()!="⚠"].copy()
    if col_p: df=df[df[col_p].astype(str).str.strip()!=""].copy()

    def _norm(s):
        s=str(s).strip(); s=_ud.normalize("NFKD",s)
        s="".join(ch for ch in s if not _ud.combining(ch))
        return " ".join(s.lower().split())

    _SEM={"","-","—","unica","única","unidade","único","unico","nan","none"}
    _ALIAS={
        _norm("Pega Areia"):            [_norm("Tapete pega areia")],
        _norm("Arranhador de Braço"):   [_norm("Arranhador Braço")],
        _norm("Arranhador de Braço Simples"):[_norm("Arranhador Braço Simples")],
        _norm("Arranhador de Madeira"): [_norm("Arranhador Madeira")],
        _norm("Brinquedo Interativo"):  [_norm("Bririnquedo interativo")],
        _norm("Arranhador Túnel"):      [_norm("Tunel para gatos")],
        _norm("Sanitario e Pega Areia"):[_norm("Sanitário + PA"),_norm("Sanitario + PA")],
        _norm("Alimentador Automático"):[_norm("Comedouro Automatico"),_norm("Alimentador Automatico")],
        _norm("Banheiro Inteligente"):  [_norm("Sanitário Automatico"),_norm("Sanitario Automatico")],
        _norm("Pato Interativo"):       [_norm("Pato Ducky")],
        _norm("Piso Carpete"):          [_norm("Piso Carpete 50x50")],
        _norm("Arranhador Cara de Gato"):[_norm("Arranhador cara gato"),_norm("Arranhador Adesivo Gato")],
        _norm("Camisa Dryfit"):         [_norm("Camisa Dry Fit")],
        _norm("Cama Suspensa"):         [_norm("Caminha Rede")],
        _norm("Poste Arranhador"):      [_norm("Arranhador Poste")],
        _norm("Playground"):            [_norm("Playground")],
        _norm("Playgrounds 15MM"):      [_norm("Playground")],
    }

    _lk={}
    if tabela_custo_fech is not None and not tabela_custo_fech.empty:
        for _,_r in tabela_custo_fech.iterrows():
            _pk=_norm(_r["PRODUTO"]); _ck=_norm(_r.get("COR","")); _tk=_norm(_r.get("TAMANHO",""))
            _v=_r["CUSTO"]
            if _pk and pd.notna(_v) and float(_v)>0:
                _ck2=""if _ck in _SEM else _ck; _tk2=""if _tk in _SEM else _tk
                _lk[(_pk,_ck2,_tk2)]=float(_v)

    def _custo(prod,cor="",tam=""):
        pk=_norm(prod); ck=_norm(cor); ck=""if ck in _SEM else ck
        tk=_norm(tam);  tk=""if tk in _SEM else tk
        for cand in [pk]+_ALIAS.get(pk,[]):
            v=_lk.get((cand,ck,tk))
            if v: return v
        return 0.0

    _SEM_FILTRO={"","-","—","unica","única","nan","none"}

    def _soma(prod,col_f1=None,v1=None,col_f2=None,v2=None):
        mask=(df[col_p].astype(str).str.strip()==str(prod).strip() if col_p else pd.Series([True]*len(df)))
        if col_f1 and v1 is not None and col_f1 in df.columns:
            if str(v1).strip().lower() not in _SEM_FILTRO:
                mask&=df[col_f1].astype(str).str.strip().str.lower()==str(v1).strip().lower()
        if col_f2 and v2 is not None and col_f2 in df.columns:
            if str(v2).strip().lower() not in _SEM_FILTRO:
                mask&=df[col_f2].astype(str).str.strip().str.lower()==str(v2).strip().lower()
        sub=df[mask]
        qty=sub["_qty"].sum(); rec=sub["_rec"].sum()
        cor_v=v1 if col_f1==col_c else (v2 if col_f2==col_c else "")
        tam_v=v1 if col_f1==col_t else (v2 if col_f2==col_t else "")
        return qty,rec,_custo(prod,cor=cor_v or "",tam=tam_v or "")

    NOBRD=Border()
    L=Alignment(horizontal="left",vertical="center")
    C=Alignment(horizontal="center",vertical="center")
    R=Alignment(horizontal="right",vertical="center")
    FMT_M='R$ #,##0.00'; FMT_Z='R$ #,##0.00_);"-"'; FMT_I='#,##0'; FMT_P='0%'
    CIANO="00B0F0"; VERDE="92D050"; AMARELO="FFFF00"

    def _f(bold=False,color="000000",size=11): return Font(name="Calibri",bold=bold,color=color,size=size)
    def _bg(h): return PatternFill("solid",start_color=h)
    fn=_f()

    for cl,wd in [("A",8.89),("B",22.33),("C",56.33),("D",8.55),("E",12.44),
                  ("F",14.66),("G",12.78),("H",12.89),("I",15.78),("J",14.0),("K",4.44)]:
        ws.column_dimensions[cl].width=wd

    def _w(row,col,val,font=None,fill=None,align=None,fmt=None):
        cell=ws.cell(row=row,column=col,value=val)
        cell.border=NOBRD
        if font:  cell.font=font
        if fill:  cell.fill=fill
        if align: cell.alignment=align
        if fmt:   cell.number_format=fmt
        return cell

    _t = Side(style="thin",   color="BFBFBF")
    _m = Side(style="medium", color="595959")
    _n = Side()

    def _dado(lin, val_b, val_c, qty, rec, custo, is_first=False, is_last=False, sep_tam=False):
        _sep = Side(style="thin", color="595959")
        top    = _m if is_first else _t
        bot    = _m if is_last  else (_sep if sep_tam else _t)
        brd_bc = Border(top=top, bottom=bot, left=_m, right=_n)
        brd_cc = Border(top=top, bottom=bot, left=_n, right=_n)
        brd_dk = Border(top=top, bottom=bot, left=_t, right=_t)
        brd_k  = Border(top=top, bottom=bot, left=_t, right=_m)

        cell_b = ws.cell(row=lin, column=2)
        if val_b is not None:
            cell_b.value = val_b; cell_b.font = fn; cell_b.alignment = R
        cell_b.border = brd_bc

        cell_c = ws.cell(row=lin, column=3)
        cell_c.value = str(val_c) if val_c is not None else ""
        cell_c.font = fn; cell_c.alignment = L; cell_c.border = brd_cc

        _w(lin,4,int(qty),fn,align=R,fmt=FMT_I); ws.cell(lin,4).border=brd_dk
        _w(lin,5,f"=IFERROR(F{lin}/D{lin},0)",fn,align=R,fmt=FMT_M); ws.cell(lin,5).border=brd_dk
        _w(lin,6,round(rec,2) if rec else 0,fn,align=R,fmt=FMT_Z); ws.cell(lin,6).border=brd_dk
        _w(lin,7,round(custo,2) if custo else 0,fn,align=R,fmt=FMT_Z); ws.cell(lin,7).border=brd_dk
        _w(lin,8,f"=G{lin}*D{lin}",fn,align=R,fmt=FMT_M); ws.cell(lin,8).border=brd_dk
        _w(lin,9,f"=IFERROR(E{lin}-G{lin},0)",fn,align=R,fmt=FMT_Z); ws.cell(lin,9).border=brd_dk
        _w(lin,10,f"=F{lin}-H{lin}",fn,align=R,fmt=FMT_M); ws.cell(lin,10).border=brd_dk
        _w(lin,11,f'=IFERROR(IF(F{lin}=0,"",J{lin}/F{lin}),"")',fn,align=R,fmt=FMT_P)
        ws.cell(lin,11).border=brd_k
        ws.row_dimensions[lin].height=15

    def _total(lin,r0,r1):
        ws.row_dimensions[lin].height=15
        brd_tot_b  = Border(bottom=_m, left=_m, right=_n)
        brd_tot_c  = Border(bottom=_m, left=_n, right=_n)
        brd_tot_dk = Border(top=_t,   bottom=_m, left=_t, right=_t)
        brd_tot_k  = Border(top=_t,   bottom=_m, left=_t, right=_m)
        ws.cell(lin,2).border=brd_tot_b
        ws.cell(lin,3).border=brd_tot_c
        _w(lin,4,f"=SUM(D{r0}:D{r1})",fn,align=R,fmt=FMT_I); ws.cell(lin,4).border=brd_tot_dk
        ws.cell(lin,5).border=brd_tot_dk
        _w(lin,6,f"=SUM(F{r0}:F{r1})",fn,fill=_bg(CIANO),align=R,fmt=FMT_M); ws.cell(lin,6).border=brd_tot_dk
        ws.cell(lin,7).border=brd_tot_dk
        _w(lin,8,f"=SUM(H{r0}:H{r1})",fn,align=R,fmt=FMT_M); ws.cell(lin,8).border=brd_tot_dk
        ws.cell(lin,9).border=brd_tot_dk
        _w(lin,10,f"=SUM(J{r0}:J{r1})",fn,fill=_bg(AMARELO),align=R,fmt=FMT_M); ws.cell(lin,10).border=brd_tot_dk
        _w(lin,11,f'=IFERROR(IF(F{lin}=0,"",J{lin}/F{lin}),"")',fn,align=R,fmt=FMT_P)
        ws.cell(lin,11).border=brd_tot_k

    _TAM_ORD={"P":1,"M":2,"G":3,"GG":4,"XGG":5}
    def _sort_tam(t):
        t=str(t).strip()
        if t in _TAM_ORD: return (0,_TAM_ORD[t],t)
        try:
            parts=t.replace("x","X").split("X")
            return (1,int(parts[0]),t)
        except: return (2,0,t)

    ESTRUTURA=[
        ("Pega Areia","Cor","COR",["Preto","Azul","Amarelo","Rosa"],"Pega Areia"),
        ("Tapete Lavável","Modelo","TAM",["P","M","G"],"Tapete Lavável"),
        ("Comedouro","Cor","COR",["Marrom","Azul","Vermelho","Amarelo","Laranja","Verde","Roxo"],None),
        ("Arranhador de Braço","Modelo","TAM_COR_MAP",
         [("70X50","P", ["Azul","Bege","Cinza"]),("100X50","M",["Azul","Bege","Cinza"]),
          ("120X50","G",["Azul","Bege","Cinza"])],"Arranhador de Braço"),
        ("Arranhador de Braço Simples","Modelo","TAM_COR_MAP",
         [("70X50","P", ["Azul","Bege","Cinza"]),("100X50","M",["Azul","Bege","Cinza"]),
          ("120X50","G",["Azul","Bege","Cinza"])],"Arranhador de Braço Simples"),
        ("Arranhador Adesivo","Modelo","TAM_COR",
         [("50x30",["Cinza","Grafite","Preto","Marrom","Bege","Azul"]),
          ("70x50",["Cinza","Grafite","Preto","Marrom","Bege","Azul"]),
          ("100x50",["Cinza","Grafite","Preto","Marrom","Bege","Azul"]),
          ("120x50",["Cinza","Grafite","Preto","Marrom","Bege","Azul"]),
          ("200x50",["Cinza","Grafite","Preto","Marrom","Bege","Azul"])],None),
        ("Arranhador de Madeira","Modelo","COR",["60cm","Cinza","Bege","Azul"],"Arranhador de Madeira"),
        ("Casinha Toca","Cor","COR",["Bege","Cinza","Caramelo","Verde","Azul","Chumbo"],None),
        ("Capa para Carro","Cor","COR",["Preto"],None),
        ("Camisa Dryfit","Modelo","TAM_COR",
         [("P",["Preto","Cinza","Verde","Branco"]),("M",["Preto","Cinza","Verde","Branco"]),
          ("G",["Preto","Cinza","Verde","Branco"]),("GG",["Preto","Cinza","Verde","Branco"]),
          ("XGG",["Preto","Cinza","Verde","Branco"])],"Camisa Dryfit"),
        ("Casinhas / Caminhas","Modelo","GRUPO_B",
         [("Casinha 01","Bege"),("Casinha 01","Cinza"),("Casinha 01","Azul"),
          ("Caminha 02","Bege"),("Caminha 02","Cinza"),("Caminha 02","Azul"),
          ("Casinha 03","Bege"),("Casinha 03","Cinza"),("Casinha 03","Azul"),
          ("Casinha 04","Bege"),("Casinha 04","Cinza"),("Casinha 04","Azul"),
          ("Casinha 05","Bege"),("Casinha 05","Cinza"),("Casinha 05","Azul"),
          ("Casinha 06","Bege"),("Casinha 06","Cinza"),("Casinha 06","Azul"),
          ("Casinha 07","Bege"),("Casinha 07","Cinza"),("Casinha 07","Azul"),
          ("Casinha 08","Bege"),("Casinha 08","Cinza"),("Casinha 08","Azul"),
          ("Casinha 09","Bege"),("Casinha 09","Cinza"),("Casinha 09","Azul"),
          ("Casinha 10","Bege"),("Casinha 10","Cinza"),("Casinha 10","Azul"),
          ("Casinha 11","Bege"),("Casinha 11","Cinza"),("Casinha 11","Azul"),
          ("Casinha 12","Bege"),("Casinha 12","Cinza"),("Casinha 12","Azul"),
          ("Casinha 13","Bege"),("Casinha 13","Cinza"),("Casinha 13","Azul"),
          ("Casinha 14","Bege"),("Casinha 14","Cinza"),("Casinha 14","Azul"),
          ("Casinha 15","Bege"),("Casinha 15","Cinza"),("Casinha 15","Azul"),
          ("Casinha 16","Bege"),("Casinha 16","Cinza"),("Casinha 16","Azul")],None),
        ("Playground","Modelo","COR",["-"],"Playground"),
        ("Chuchups","Modelo","COR",["-"],None),
        ("Presilhas Aspirais","Modelo","COR",["Branco"],None),
        ("Piso Carpete","Modelo","COR",["Cinza","Grafite","Preto","Marrom","Bege","Azul"],"Piso Carpete"),
        ("Sanitário","Modelo","COR",["Bege","Cinza","Preto","Verde"],"Sanitário"),
        ("Sanitario e Pega Areia","Cor","COR",["Bege","Cinza","Preto","Verde"],"Sanitario e Pega Areia"),
        ("Cama Suspensa","Modelo","COR",["Bege","Cinza","Caramelo","Verde","Azul","Chumbo"],"Cama Suspensa"),
        ("Arranhador Túnel","Modelo","COR",["Bege","Cinza","Caramelo","Verde","Azul","Chumbo"],"Arranhador Túnel"),
        ("Degrau","Modelo","COR",["Bege","Cinza","Marrom","Verde","Azul","Chumbo"],None),
        ("Caminha Comeia","Modelo","COR",["Bege","Cinza","Caramelo","Verde","Azul","Chumbo"],None),
        ("Alimentador Automático","Modelo","COR",["Branco"],"Alimentador Automático"),
        ("Brinquedo Interativo","Modelo","COR",["Azul","Bege","Cinza","Grafite","Marrom","Preto"],"Brinquedo Interativo"),
        ("Arranhadores","Cor","GRUPO_B",
         [("Arranhador Rampa","Bege"),("Arranhador Rampa","Cinza"),("Arranhador Rampa","Azul"),
          ("Arranhador Parede","Sisal"),("Arranhador Parede","Bege"),("Arranhador Parede","Cinza"),
          ("Arranhador Parede","Preto"),("Arranhador Parede","Marrom"),
          ("Arranhador Parede","Azul"),("Arranhador Parede","Grafite"),
          ("Arranhador Cara de Gato","Bege"),("Arranhador Cara de Gato","Cinza"),
          ("Arranhador Cara de Gato","Preto"),("Arranhador Cara de Gato","Verde"),
          ("Arranhador Cara de Gato","Azul"),("Arranhador Cara de Gato","Chumbo"),
          ("Poste Arranhador","Sisal")],None),
        ("Arranhador Adesivo Gato","Modelo","COR",["Azul","Bege","Cinza","Grafite","Marrom","Preto"],None),
        ("Pato Interativo","Modelo","COR",["Amarelo"],"Pato Interativo"),
        ("Banheiro Inteligente","Modelo","COR",["-"],"Banheiro Inteligente"),
        ("Comedouro Inteligente","Modelo","COR",["Branco"],None),
        ("Playgrounds 15MM","Modelo","GRUPO_B",
         [("Playground 02","-"),("Playground 03","-"),("Playground 04","-"),
          ("Playground 05","-"),("Playground 06","-")],None),
        ("Nicho Grande","Modelo","COR",["Única"],None),
        ("FullSellers","Modelo","COR",
         ["Bolsa Caixa Transporte Aves Calopsita Periquito Pássaros Top Cor Azul",
          "Gaiola Hamster Tubo Super Luxo 3 Andares Porquinho Da India",
          "Gaiola Hamster 2 Andares Porquinho Da Inda Chinchila"],None),
    ]

    ws.row_dimensions[3].height=15
    _w(3,2,"MERCADO LIVRE",fn,fill=_bg(AMARELO),align=L)
    linha=4; linhas_total=[]

    for nome_exib,tipo_exib,modo,itens,nome_bf in ESTRUTURA:
        prod_bf=nome_bf if nome_bf else nome_exib
        ws.row_dimensions[linha].height=15
        brd_hdr_b  = Border(top=_m, bottom=_t, left=_m, right=_n)
        brd_hdr_c  = Border(top=_m, bottom=_t, left=_n, right=_n)
        brd_hdr_dk = Border(top=_m, bottom=_t, left=_t, right=_t)
        brd_hdr_k  = Border(top=_m, bottom=_t, left=_t, right=_m)
        _w(linha,2,nome_exib,fn,align=L); ws.cell(linha,2).border=brd_hdr_b
        _w(linha,3,tipo_exib,fn,align=L); ws.cell(linha,3).border=brd_hdr_c
        for col_n,lbl in [(4,"Unidades"),(5,"Valor Unitario"),(6,"Valor de entrada"),
                          (7,"Custo Unitario"),(8,"Custo Total"),
                          (9,"Lucro por unidade"),(10,"Lucro total")]:
            brd = brd_hdr_k if col_n==10 else brd_hdr_dk
            _w(linha,col_n,lbl,fn,align=C); ws.cell(linha,col_n).border=brd
        ws.cell(linha,11).border=Border(top=_m, bottom=_t, left=_t, right=_m)
        linha+=1; r_ini=linha

        if modo=="COR":
            itens_ord=sorted(itens,key=str.lower); n_itens=len(itens_ord)
            for i_cor,cor in enumerate(itens_ord):
                qty,rec,cu=_soma(prod_bf,col_c,cor)
                _dado(linha,None,cor,qty,rec,cu,is_first=(i_cor==0),is_last=(i_cor==n_itens-1))
                linha+=1
        elif modo=="TAM":
            n_itens=len(itens)
            for i_tam,tam in enumerate(itens):
                qty,rec,cu=_soma(prod_bf,col_t,tam)
                _dado(linha,None,tam,qty,rec,cu,is_first=(i_tam==0),is_last=(i_tam==n_itens-1))
                linha+=1
        elif modo=="TAM_COR":
            todas_linhas=[(tam,cor) for tam,cores in itens for cor in sorted(cores,key=str.lower)]
            n_total=len(todas_linhas); i_global=0
            for i_tam,(tam,cores) in enumerate(itens):
                cores_ord=sorted(cores,key=str.lower); ultimo_tam=(i_tam==len(itens)-1)
                for i_cor,cor in enumerate(cores_ord):
                    ultima_cor=(i_cor==len(cores_ord)-1)
                    is_last_block=(i_global==n_total-1); is_sep_tam=ultima_cor and not ultimo_tam
                    qty,rec,cu=_soma(prod_bf,col_t,tam,col_c,cor)
                    _dado(linha,tam,cor,qty,rec,cu,is_first=(i_global==0),is_last=is_last_block,sep_tam=is_sep_tam)
                    i_global+=1; linha+=1
        elif modo=="TAM_COR_MAP":
            todas_linhas=[(r,t,cor) for r,t,cores in itens for cor in sorted(cores,key=str.lower)]
            n_total=len(todas_linhas); i_global=0
            for i_tam,(rotulo,tam_bf,cores) in enumerate(itens):
                cores_ord=sorted(cores,key=str.lower); ultimo_tam=(i_tam==len(itens)-1)
                for i_cor,cor in enumerate(cores_ord):
                    ultima_cor=(i_cor==len(cores_ord)-1)
                    is_last_block=(i_global==n_total-1); is_sep_tam=ultima_cor and not ultimo_tam
                    qty,rec,cu=_soma(prod_bf,col_t,tam_bf,col_c,cor)
                    if cu==0: cu=_custo(prod_bf,cor=cor,tam=rotulo)
                    _dado(linha,rotulo,cor,qty,rec,cu,is_first=(i_global==0),is_last=is_last_block,sep_tam=is_sep_tam)
                    i_global+=1; linha+=1
        elif modo=="GRUPO_B":
            itens_validos=[(sb,sc) for sb,sc in itens if sb]; n_itens=len(itens_validos)
            for i_item,(sub_b,sub_c) in enumerate(itens_validos):
                sub_prod=str(sub_b)
                if sub_c: qty,rec,cu=_soma(sub_prod,col_c,sub_c)
                else: qty,rec,cu=_soma(sub_prod); cu=_custo(sub_prod)
                _dado(linha,sub_b,sub_c,qty,rec,cu,is_first=(i_item==0),is_last=(i_item==n_itens-1))
                linha+=1

        r_fim=linha-1
        _total(linha,r_ini,r_fim)
        linhas_total.append(linha); linha+=1
        ws.row_dimensions[linha].height=10; linha+=1

    ws.row_dimensions[linha].height=7; linha+=1
    ws.row_dimensions[linha].height=15
    refs_d="+".join(f"D{r}" for r in linhas_total)
    refs_f="+".join(f"F{r}" for r in linhas_total)
    refs_h="+".join(f"H{r}" for r in linhas_total)
    refs_j="+".join(f"J{r}" for r in linhas_total)
    _w(linha,2,"TOTAL",fn,align=L)
    _w(linha,4,f"={refs_d}",fn,align=R,fmt=FMT_I)
    _w(linha,6,f"={refs_f}",fn,align=R,fmt=FMT_M)
    _w(linha,8,f"={refs_h}",fn,align=R,fmt=FMT_M)
    _w(linha,10,f"={refs_j}",fn,align=R,fmt=FMT_M)
    _w(linha,11,f'=IFERROR(IF(F{linha}=0,"",J{linha}/F{linha}),"")',fn,align=R,fmt=FMT_P)
    ws.freeze_panes="A4"

# =============================================================================
# ABA PRODUTOS
# =============================================================================

def _escrever_painel_produtos(wb, base_final_df):
    ws = wb.create_sheet("Produtos")
    df = base_final_df.copy()

    def _f(bold=False, color="1A1A1A", size=10, italic=False):
        return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)
    def _bg(h): return PatternFill("solid", start_color=h)
    thin = Side(style="thin", color="D9D9D9")
    BRD  = Border(left=thin, right=thin, top=thin, bottom=thin)
    L = Alignment(horizontal="left",   vertical="center")
    C = Alignment(horizontal="center", vertical="center")
    R = Alignment(horizontal="right",  vertical="center")
    FMT = 'R$ #,##0.00'; FPCT = '0.0%'; FINT = '#,##0'

    NAVY = "1F3864"; BLUE = "2E75B6"; LBLUE = "BDD7EE"
    GREY = "F2F2F2"; WHITE = "FFFFFF"; GREEN = "375623"

    CANAIS = ["Mercado Livre", "Shopee", "Magalu", "Amazon", "Shein", "Site Próprio", "TOTAL"]
    COR_CANAL = {
        "Mercado Livre": BLUE, "Shopee": "F47920", "Magalu": "0070C0",
        "Amazon": "FF9900", "Shein": "E83E8C", "Site Próprio": GREEN, "TOTAL": NAVY,
    }

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 28
    for i in range(len(CANAIS) + 1):
        ws.column_dimensions[get_column_letter(3 + i)].width = 15

    def _col(df, *names):
        for n in names:
            m = next((x for x in df.columns if str(x).strip().lower() == n.lower()), None)
            if m: return m
        return None

    col_produto = _col(df, "PRODUTO")
    col_receita = _col(df, "Receita por produtos (BRL)")
    col_qty     = _col(df, "TOTAL UNIDADES")
    col_unid    = _col(df, "Unidades")

    def _num(c):
        return pd.to_numeric(df[c], errors="coerce").fillna(0) if c else pd.Series([0.0]*len(df))

    df["_rec"] = _num(col_receita)
    df["_qty"] = _num(col_qty) if col_qty else _num(col_unid)

    if col_produto:
        por_prod = (df[df[col_produto].astype(str).str.strip() != ""]
                    .groupby(col_produto)
                    .agg(fat=("_rec","sum"), qty=("_qty","sum"))
                    .sort_values("fat", ascending=False)
                    .reset_index())
    else:
        por_prod = pd.DataFrame()

    ORDEM = [
        "Pega Areia","Tapete Lavável","Comedouro","Arranhador de Braço",
        "Arranhador de Braço Simples","Arranhador Adesivo","Arranhador de Madeira",
        "Casinha Toca","Capa para Carro","Camisa Dryfit","Casinhas / Caminhas",
        "Playground","Chuchups","Presilhas Aspirais","Piso Carpete","Sanitário",
        "Sanitario e Pega Areia","Caminha Rede","Arranhador Túnel","Degrau",
        "Caminha Comeia","Comedouro Automático","Brinquedo Interativo","Arranhadores",
        "Pato Interativo","Sanitário Automático","Comedouro Inteligente","FullSellers",
    ]
    produtos_bf = set(por_prod[col_produto].astype(str).str.strip().tolist()) if len(por_prod) else set()
    extras      = sorted([p for p in produtos_bf - set(ORDEM) if p not in ("","nan")])
    ordem_final = ORDEM + extras

    n_cols = 2 + len(CANAIS) + 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1, value="PRODUTOS  —  Faturamento por Canal de Venda")
    c.font = _f(bold=True, color="FFFFFF", size=13)
    c.fill = _bg(NAVY); c.alignment = C; ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    c = ws.cell(row=2, column=1,
                value="Mercado Livre: dados reais  |  Demais canais: estrutura preparada para implementação futura")
    c.font = _f(size=9, color="7F6000")
    c.fill = _bg("FFF2CC"); c.alignment = C; ws.row_dimensions[2].height = 14

    c = ws.cell(row=3, column=2, value="Produto")
    c.font = _f(bold=True, color="FFFFFF", size=9)
    c.fill = _bg(NAVY); c.alignment = C; c.border = BRD
    for i, canal in enumerate(CANAIS):
        col = 3 + i
        c = ws.cell(row=3, column=col, value=canal)
        c.font = _f(bold=True, color="FFFFFF", size=9)
        c.fill = _bg(COR_CANAL.get(canal, NAVY)); c.alignment = C; c.border = BRD
    c = ws.cell(row=3, column=3+len(CANAIS), value="% ML s/ Total")
    c.font = _f(bold=True, color="FFFFFF", size=9)
    c.fill = _bg(BLUE); c.alignment = C; c.border = BRD
    ws.row_dimensions[3].height = 16

    linha = 4; linhas_fat = []; i_alt = 0
    for prod in ordem_final:
        bg = _bg(GREY if i_alt % 2 == 0 else WHITE); i_alt += 1
        if len(por_prod) > 0 and col_produto:
            rp = por_prod[por_prod[col_produto].astype(str).str.strip() == prod]
            fat_ml = rp["fat"].sum() if len(rp) else 0
        else:
            fat_ml = 0

        c = ws.cell(row=linha, column=2, value=prod)
        c.font = _f(size=9); c.fill = bg; c.alignment = L; c.border = BRD

        for i, canal in enumerate(CANAIS):
            col = 3 + i
            if canal == "Mercado Livre":
                val = fat_ml; cf = _f(size=9); cf2 = bg
            elif canal == "TOTAL":
                refs = "+".join(get_column_letter(3+j)+str(linha) for j in range(len(CANAIS)-1))
                val = f"={refs}"; cf = _f(bold=True, size=9); cf2 = _bg(LBLUE)
            else:
                val = None; cf = _f(size=9, color="9E9E9E"); cf2 = bg
            c = ws.cell(row=linha, column=col, value=val)
            c.font = cf; c.fill = cf2; c.alignment = R; c.number_format = FMT; c.border = BRD

        col_tot = 3 + len(CANAIS) - 1
        col_pct = 3 + len(CANAIS)
        c = ws.cell(row=linha, column=col_pct,
                    value=f'=IFERROR({get_column_letter(3)}{linha}/'
                           f'{get_column_letter(col_tot)}{linha},"")')
        c.font = _f(size=9, italic=True); c.fill = bg
        c.alignment = R; c.number_format = FPCT; c.border = BRD
        ws.row_dimensions[linha].height = 14
        linhas_fat.append(linha); linha += 1

    c = ws.cell(row=linha, column=2, value="TOTAL GERAL")
    c.font = _f(bold=True, color="FFFFFF", size=10)
    c.fill = _bg(NAVY); c.alignment = L; c.border = BRD
    for i, canal in enumerate(CANAIS):
        col = 3 + i
        cl  = get_column_letter(col)
        c = ws.cell(row=linha, column=col,
                    value=f"=SUM({cl}{linhas_fat[0]}:{cl}{linhas_fat[-1]})")
        c.font = _f(bold=True, color="FFFFFF", size=10)
        c.fill = _bg(COR_CANAL.get(canal, NAVY)); c.alignment = R
        c.number_format = FMT; c.border = BRD
    ws.cell(row=linha, column=3+len(CANAIS)).fill = _bg(NAVY)
    ws.cell(row=linha, column=3+len(CANAIS)).border = BRD
    ws.row_dimensions[linha].height = 18
    ws.freeze_panes = "C4"

# =============================================================================
# ORQUESTRADOR INTERNO
# =============================================================================

def _escrever_paineis(wb, base_final_df, tabela_custo=None, tabela_custo_fech=None):
    _escrever_fechamento(wb, base_final_df, tabela_custo_fech=tabela_custo_fech)
    _escrever_painel_produtos(wb, base_final_df)

# =============================================================================
# INTERFACE PÚBLICA
# =============================================================================

def processar(arquivos_input: list, tabela_bytes: bytes):
    """
    arquivos_input : list of (nome: str, conteudo: bytes)
    tabela_bytes   : conteúdo de TABELA_AUXILIAR.xlsx em bytes
    Retorna        : (xlsx_bytes: bytes, logs: list[str])
    """
    log = SimpleLog()

    ricapet_rows, ricapet_cols = [], []
    thapets_rows, thapets_cols = [], []

    for nome, conteudo in arquivos_input:
        empresa = identificar_empresa(nome)
        if empresa is None:
            log.warning(f"Ignorado (empresa não identificada pelo nome do arquivo): {nome}")
            continue
        rows, cols = ler_arquivo(conteudo, nome, empresa, log)
        if empresa == "Ricapet":
            ricapet_rows.extend(rows)
            if not ricapet_cols:
                ricapet_cols = cols
        else:
            thapets_rows.extend(rows)
            if not thapets_cols:
                thapets_cols = cols

    if not ricapet_rows and not thapets_rows:
        raise ValueError(
            "Nenhum dado foi lido. Verifique se os arquivos enviados são relatórios "
            "do Mercado Livre da Ricapet ou Thapets (os nomes devem conter "
            f"'{CHAVE_RICAPET}' ou '{CHAVE_THAPETS}')."
        )

    if not ricapet_cols:
        ricapet_cols = thapets_cols
    if not thapets_cols:
        thapets_cols = ricapet_cols

    rows, all_cols = unificar(ricapet_rows, ricapet_cols, thapets_rows, thapets_cols, log)
    groups = construir_grupos(rows)
    log.info(f"Grupos mãe/filhos identificados: {len(groups)}")
    rows = processar_grupos(rows, groups, log)
    erros = gerar_erros(rows, groups, log)

    log.info("Carregando TABELA_AUXILIAR...")
    tabela_produtos, tabela_status, tabela_custo, tabela_custo_fech = carregar_auxiliares(tabela_bytes, log)

    log.info("Criando Base Final...")
    base_final_df = criar_base_final(rows, all_cols, log)

    log.info("Preenchendo colunas calculadas...")
    base_final_df = preencher_colunas(
        base_final_df, tabela_produtos, tabela_status, log, tabela_custo=tabela_custo
    )

    log.info("Gerando arquivo Excel...")
    output = io.BytesIO()
    escrever_xlsx(rows, all_cols, erros, base_final_df, output, log, tabela_custo, tabela_custo_fech)
    output.seek(0)

    log.info(f"Concluído: {len(base_final_df)} linhas na Base Final, {len(erros)} divergência(s)")
    return output.read(), log.msgs
